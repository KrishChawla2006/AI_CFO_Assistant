import os
import re
import warnings
import logging
import json
import time
import io
from datetime import datetime, timedelta
from contextlib import contextmanager, redirect_stdout, redirect_stderr
from concurrent.futures import ThreadPoolExecutor, as_completed

from flask import Flask, render_template, request, jsonify, session, send_file, redirect, url_for, flash,Response,stream_with_context
from flask_session import Session
from flask_sqlalchemy import SQLAlchemy
from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, session, flash
from dotenv import load_dotenv
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
from reportlab.lib.units import inch
import pandas as pd
import numpy as np

import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio

import pdfplumber
try:
    import fitz  
except ImportError:
    fitz = None
    logging.warning("PyMuPDF (fitz) not found. PDF processing will be limited.")

from PIL import Image
try:
    import pytesseract
except ImportError:
    pytesseract = None
    logging.warning("Tesseract OCR (pytesseract) not found. Image-based text extraction will be limited.")

from openpyxl import load_workbook

# AI and ML libraries
from textblob import TextBlob
from sklearn.ensemble import RandomForestRegressor, GradientBoostingRegressor
from sklearn.linear_model import LinearRegression, Ridge
from sklearn.preprocessing import StandardScaler
from sklearn.metrics import mean_absolute_error
 
load_dotenv()

app = Flask(__name__)

app.secret_key = os.getenv("FLASK_SECRET")

 
users = {}
 
try:
    from prophet import Prophet
    PROPHET_AVAILABLE = True
except ImportError:
    Prophet = None
    PROPHET_AVAILABLE = False
    logging.warning("Prophet not found. Time series forecasting will use sklearn models only.")



try:
    from reportlab.lib.pagesizes import letter
    from reportlab.pdfgen import canvas
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    logging.error("ReportLab not installed. PDF export will not function.")
    Paragraph, Spacer, Table, TableStyle, getSampleStyle_Sheet, ParagraphStyle, SimpleDocTemplate = (
        None, None, None, None, None, None, None
    )
    REPORTLAB_AVAILABLE = False

 
try:
    import requests
    OLLAMA_AVAILABLE = True
except ImportError:
    requests = None
    OLLAMA_AVAILABLE = False
    logging.warning("Requests library not found. Ollama integration will be disabled.")

 
try:
    import Levenshtein
except ImportError:
    logging.warning("python-Levenshtein not found. Falling back to a slower pure-Python Levenshtein implementation.")
 
    def levenshtein_distance(s1, s2):
        if len(s1) < len(s2):
            return levenshtein_distance(s2, s1)
        if len(s2) == 0:
            return len(s1)
        previous_row = list(range(len(s2) + 1))
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        return previous_row[-1]
    Levenshtein = type('Levenshtein', (object,), {'distance': staticmethod(levenshtein_distance)})

from dotenv import load_dotenv

 
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

 
load_dotenv()

logging.info(f"Current working directory: {os.getcwd()}")
logging.info(f"Files in directory: {os.listdir('.')}")
logging.info(f".env file exists: {os.path.exists('.env')}")

SECRET_KEY = os.getenv("SECRET_KEY", os.urandom(24))

 
OLLAMA_API_BASE_URL = os.getenv("OLLAMA_API_BASE_URL", "http://localhost:11434")
 
 
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama2:7b-chat")  
USE_OLLAMA = OLLAMA_AVAILABLE

if USE_OLLAMA:
    try:
        test_response = requests.get(f"{OLLAMA_API_BASE_URL}/api/tags")
        test_response.raise_for_status()
        models = [m['name'] for m in test_response.json().get('models', [])]
        if OLLAMA_MODEL not in models:
            logging.warning(f"Configured OLLAMA_MODEL '{OLLAMA_MODEL}' not found. Available models: {', '.join(models)}. Ollama features might not work as expected. Please download it using 'ollama run {OLLAMA_MODEL}'.")
        logging.info(f"Ollama API connected successfully. Using model: {OLLAMA_MODEL}")
    except requests.exceptions.ConnectionError:
        logging.error(f"Could not connect to Ollama API at {OLLAMA_API_BASE_URL}. Is Ollama running?")
        USE_OLLAMA = False
    except requests.exceptions.RequestException as e:
        logging.error(f"Error communicating with Ollama API: {e}")
        USE_OLLAMA = False
else:
    logging.warning("Requests library not available or Ollama explicitly disabled. Ollama integration will be bypassed.")


 
app.config['SECRET_KEY'] = SECRET_KEY
app.config['SESSION_TYPE'] = 'filesystem'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['REPORTS_FOLDER'] = 'reports'
app.config['MAX_CONTENT_LENGTH'] = 500 * 1024 * 1024
app.secret_key = 'your_super_secret_key_for_session_management'
REPORTS_FOLDER = os.path.join(os.getcwd(), 'exported_reports')
os.makedirs(REPORTS_FOLDER, exist_ok=True)
app.config['REPORTS_FOLDER'] = REPORTS_FOLDER
 
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///site.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

 
Session(app)

 
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['REPORTS_FOLDER'], exist_ok=True)

USE_GEMINI = False
logging.info("Gemini API integration has been removed as requested.")

 
warnings.filterwarnings("ignore")
logging.getLogger('pdfplumber').setLevel(logging.CRITICAL)

 
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls', 'csv'}

 
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128), nullable=False)
    company_name = db.Column(db.String(120), nullable=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def __repr__(self):
        return f'<User {self.username}>'

 
def login_required(f):
    @wraps(f)
    # @app.before_request
    def decorated_function(*args, **kwargs):
        if 'user' not in session:
            flash('You need to log in first.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

 
@contextmanager
def suppress_output():
    """Suppress stdout and stderr for noisy libraries."""
    with open(os.devnull, 'w') as devnull:
        with redirect_stdout(devnull), redirect_stderr(devnull):
            yield

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def sanitize_for_json(data):
    """Sanitize data to ensure JSON serialization compatibility"""
    if isinstance(data, dict):
        return {key: sanitize_for_json(value) for key, value in data.items()}
    elif isinstance(data, list):
        return [sanitize_for_json(item) for item in data]
    elif isinstance(data, float):
        if np.isnan(data):
            return 0.0
        elif np.isinf(data):
            return 9999999999.99 if data > 0 else -9999999999.99
        elif abs(data) > 1e15:
            return 1e15 if data > 0 else -1e15
        return round(data, 2)
    elif isinstance(data, np.integer):
        return int(data)
    elif isinstance(data, np.floating):
        return float(sanitize_for_json(float(data)))
    elif isinstance(data, pd.DataFrame):
        return data.to_dict(orient='records')
    else:
        return data

 

 
class AdvancedIngestionAgent:
    """Enhanced ingestion agent with multiple extraction methods"""

    def __init__(self):
        self.financial_keywords = {
            'company_name': ['company', 'entity', 'organization', 'group'],
            'financial_year': ['year ended', 'financial year', 'reporting period', 'as at', 'for the year'],
            'revenue': ['revenue', 'sales', 'turnover', 'income', 'gross income', 'total income', 'operating income'],
            'expenses': ['expenses', 'costs', 'expenditure', 'operating expenses', 'total expenses', 'cost of goods sold', 'cogs'],
            'profit': ['profit', 'net income', 'net profit', 'pat', 'profit after tax', 'earnings', 'ebitda', 'ebit'],
            'assets': ['total assets', 'current assets', 'fixed assets', 'assets', 'balance sheet'],
            'liabilities': ['total liabilities', 'current liabilities', 'long term liabilities', 'debt', 'borrowings'],
            'cash': ['cash', 'cash equivalents', 'cash and cash equivalents', 'liquid assets'],
            'equity': ['shareholders equity', 'equity', 'net worth', 'book value', 'retained earnings']
        }
 

    def clean_financial_value(self, value):
        """Enhanced financial value cleaning with multiple currency support and multipliers."""
        if pd.isna(value) or not isinstance(value, (str, float, int)):
            return 0.0

        s_value = str(value).strip().lower()
        if not s_value:
            return 0.0

        multiplier = 1.0

 
        if 'lakh' in s_value:
            s_value = re.sub(r'\blakhs?\b', '', s_value)
            multiplier *= 100000
        if 'crore' in s_value or 'cr' in s_value:
            s_value = re.sub(r'\b(crores?|cr)\b', '', s_value)
            multiplier *= 10000000
        if 'million' in s_value or 'mn' in s_value:
            s_value = re.sub(r'\b(millions?|mn)\b', '', s_value)
            multiplier *= 1000000
        if 'billion' in s_value or 'bn' in s_value:
            s_value = re.sub(r'\b(billions?|bn)\b', '', s_value)
            multiplier *= 1000000000
        if 'thousand' in s_value or 'k' in s_value:
            s_value = re.sub(r'\b(thousands?|k)\b', '', s_value)
            multiplier *= 1000
        
 
        s_value = re.sub(r'[₹$€£¥%]', '', s_value)
        s_value = re.sub(r'\b(inr|usd|eur|gbp)\b', '', s_value)

        is_negative = False
        if re.search(r'\(.*\)', s_value): 
            is_negative = True
            s_value = s_value.replace('(', '').replace(')', '')
        elif s_value.startswith('-'):
            is_negative = True
            s_value = s_value[1:]  

        s_value = re.sub(r'[,\s]', '', s_value) 

        try:
             
            num = float(s_value)
            return (-num if is_negative else num) * multiplier
        except ValueError:
           
            match = re.search(r'(\d+(?:\.\d+)?(?:e[+-]?\d+)?)', s_value)
            if match:
                num = float(match.group(1)) * multiplier
                return -num if is_negative else num
            return 0.0

    def _process_pdf_page(self, doc, page_num):
        """Helper function to process a single PDF page for text and OCR."""
        page = doc.load_page(page_num)
        page_text = page.get_text()
        ocr_text_for_page = ""

        if len(page_text.strip()) < 100 and pytesseract:  
            try:
                pix = page.get_pixmap(matrix=fitz.Matrix(2,2))  
                img_bytes = pix.tobytes("png")
                img = Image.open(io.BytesIO(img_bytes))
                ocr_text_for_page = pytesseract.image_to_string(img)
            except Exception as ocr_e:
                logging.debug(f"OCR failed for page {page_num+1}: {ocr_e}")
        
        return page_text, ocr_text_for_page
    


 

    def extract_from_pdf_advanced(self, file_path):
        """Advanced PDF extraction using multiple methods with parallel processing for text/OCR."""
        extracted_data = {'text': '', 'tables': [], 'images_text': ''}
        
        start_pdf_extraction = time.time()
        logging.info(f"Starting advanced PDF extraction for {os.path.basename(file_path)}")

        doc = None
        try:
            if fitz:  
                doc = fitz.open(file_path)
                num_pages = len(doc)
                
                pages_to_process_text_ocr = min(num_pages, 100) 
                
                with ThreadPoolExecutor(max_workers=min(os.cpu_count() * 2, pages_to_process_text_ocr)) as executor:
                    futures = {executor.submit(self._process_pdf_page, doc, page_num): page_num for page_num in range(pages_to_process_text_ocr)}
                    
                    page_results_ordered = [None] * pages_to_process_text_ocr
                    for future in as_completed(futures):
                        page_num = futures[future]
                        try:
                            page_text, ocr_text_for_page = future.result()
                            page_results_ordered[page_num] = (page_text, ocr_text_for_page)
                        except Exception as e:
                            logging.error(f"Error processing page {page_num+1} with PyMuPDF/OCR: {e}")
                            page_results_ordered[page_num] = ("", "")

                full_text = ""
                full_images_text = ""
                for page_text, ocr_text_for_page in page_results_ordered:
                    full_text += page_text + "\n"
                    full_images_text += ocr_text_for_page + "\n"

                extracted_data['text'] = full_text
                extracted_data['images_text'] = full_images_text
                
                logging.info(f"PyMuPDF text and OCR extraction complete in {time.time() - start_pdf_extraction:.2f}s for {pages_to_process_text_ocr} pages.")
            else: 
                logging.warning("PyMuPDF not found, falling back to pdfplumber for text extraction.")
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages[:50]:
                        page_text = page.extract_text()
                        if page_text:
                            extracted_data['text'] += page_text + "\n"
            
        except Exception as e:
            logging.error(f"PDF text extraction failed for {os.path.basename(file_path)}: {e}", exc_info=True)
            raise

        return extracted_data

    def extract_from_excel_advanced(self, file_path):
        """Advanced Excel extraction with multiple sheet handling."""
        extracted_data = {'sheets': {}, 'tables': []}

        try:
            if file_path.endswith('.xlsx'):
                workbook = load_workbook(file_path, data_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    data = []
                    for row in sheet.iter_rows(values_only=True):
                        if any(cell is not None for cell in row):
                            data.append(row)

                    if data:
                        try:
                            
                            if data and all(isinstance(c, (str, type(None))) for c in data[0]):
                                df = pd.DataFrame(data[1:], columns=data[0])
                            else:
                                df = pd.DataFrame(data)
                        except Exception as e:
                            logging.debug(f"Could not infer header for Excel sheet {sheet_name}: {e}. Using default integer headers.")
                            df = pd.DataFrame(data)
                        
                        df = df.dropna(how='all').dropna(axis=1, how='all')
                        if not df.empty:
                            extracted_data['sheets'][sheet_name] = df
                            extracted_data['tables'].append(df)
            else: # .xls files
                excel_data = pd.read_excel(file_path, sheet_name=None)
                extracted_data['sheets'] = excel_data
                for sheet_name, df in excel_data.items():
                    df = df.dropna(how='all').dropna(axis=1, how='all')
                    if not df.empty:
                        extracted_data['tables'].append(df)

            logging.info(f"Excel extraction successful for {os.path.basename(file_path)}. Found {len(extracted_data['tables'])} tables.")
        except Exception as e:
            logging.error(f"Excel extraction failed for {os.path.basename(file_path)}: {e}", exc_info=True)
            raise

        return extracted_data

    def extract_from_csv(self, file_path):
        """Extract data from CSV files."""
        extracted_data = {'text': '', 'tables': []}
        try:
            df = pd.read_csv(file_path)
            if not df.empty:
                df = df.dropna(how='all').dropna(axis=1, how='all')
                extracted_data['tables'].append(df)
                extracted_data['text'] = df.to_string(index=False, header=True)  
            logging.info(f"CSV extraction successful for {os.path.basename(file_path)}. Found {len(extracted_data['tables'])} tables.")
        except Exception as e:
            logging.error(f"CSV extraction failed for {os.path.basename(file_path)}: {e}", exc_info=True)
            raise
        return extracted_data

    def extract_metrics_traditional(self, text_content, tables):
        """Traditional regex-based metric extraction."""
        metrics = {key: 0.0 for key in self.financial_keywords.keys() if key not in ['company_name', 'financial_year']}
        metrics['company_name'] = "Unknown"
        metrics['financial_year'] = "Unknown"
        
        full_document_content = text_content.lower()

         
        year_pattern = r'(?:year\s+ended|financial\s+year|reporting\s+period|as\s+at)\s*[:\s]*(?:(?:dec|jun|mar|sep)\s*\d{2},\s*)?(\d{4}[-/]\d{2,4}|\d{4})'
        year_match = re.search(year_pattern, full_document_content, re.IGNORECASE)
        if year_match:
            metrics['financial_year'] = year_match.group(1).replace('/', '-')

         
        company_name_patterns = [
            r'(?:for\s+)(?:the\s+)?(company|entity|group|limited|ltd|inc|corp|corporation|holdings|industries|private)\s+([a-zA-Z0-9\s,&.-]+)',
            r'annual\s+report\s+of\s+([a-zA-Z0-9\s,&.-]+)',
            r'balance\s+sheet\s+of\s+([a-zA-Z0-9\s,&.-]+)',
            r'income\s+statement\s+of\s+([a-zA-Z0-9\s,&.-]+)',
            r'(?:statement\s+of\s+)\s*(?:financial\s+position|profit\s+and\s+loss)\s+for\s+([a-zA-Z0-9\s,&.-]+)'
        ]
        
        for pattern in company_name_patterns:
            company_match = re.search(pattern, full_document_content, re.IGNORECASE)
            if company_match:
             
                name_candidate = company_match.groups()[-1].strip()
            
                name_candidate = re.split(r'(\bconsolidated\b|\bstandalone\b|\bannexure\b|\breport\b|\bstatement\b)', name_candidate, 1)[0].strip()
                if name_candidate and name_candidate.lower() not in ['limited', 'ltd', 'inc', 'corp', 'company', 'group', 'entity']:
                    metrics['company_name'] = name_candidate.title()
                    break

  
        for metric, keywords in self.financial_keywords.items():
            if metric in ['company_name', 'financial_year']: continue
            keyword_pattern = '|'.join(re.escape(k) for k in keywords)
            
            pattern = rf'(?:{keyword_pattern})\s*[:\-\—]?\s*(?:(?:₹|\$|€|£)\s*)?((?:\(|-)?\d{{1,3}}(?:,\d{{3}})*(?:\.\d{{1,2}})?(?:k|m|mn|bn|lakh|crore|cr)?(?:\))?)'
            
            matches = re.finditer(pattern, full_document_content, re.IGNORECASE)
            for match in matches:
                value = self.clean_financial_value(match.group(1))
                if value != 0.0:
                    if abs(value) > abs(metrics[metric]) or metrics[metric] == 0.0:
                        metrics[metric] = value
                
 
        for df in tables: 
            if df.empty:
                continue
            
            df_processed = df.copy()
            for col in df_processed.columns:
 
                df_processed[col] = df_processed[col].apply(self.clean_financial_value)
            
            for metric, keywords in self.financial_keywords.items():
                if metric in ['company_name', 'financial_year']: continue
                
                for r_idx, row in df_processed.iterrows():
                    if any(any(keyword in str(cell_val) for keyword in keywords) for cell_val in row.values):
                        for c_idx, cell_value in enumerate(row.values):
    
                    
                            cleaned_val = cell_value  
                            if cleaned_val != 0.0:
                                if any(kw in str(df.columns[c_idx]).lower() for kw in ['amount', 'value', 'total', 'net', '20', 'fy']):
                                    if abs(cleaned_val) > abs(metrics[metric]) or metrics[metric] == 0.0:
                                        metrics[metric] = cleaned_val
                                        break
                        if metrics[metric] != 0.0:
                            break
        
        logging.info("Traditional metric extraction successful.")
        return metrics

    def _calculate_derived_metrics(self, metrics):
        """Calculate advanced derived financial metrics, handling potential division by zero."""
        for key in ['revenue', 'expenses', 'profit', 'assets', 'liabilities', 'cash', 'equity']:
            metrics[key] = metrics.get(key, 0.0)

        metrics['profitability'] = (metrics['profit'] / metrics['revenue'] * 100) if metrics['revenue'] > 0 else 0.0
        metrics['expense_ratio'] = (metrics['expenses'] / metrics['revenue'] * 100) if metrics['revenue'] > 0 else 0.0

        metrics['debt_to_equity'] = (metrics['liabilities'] / metrics['equity']) if metrics['equity'] > 0 else 0.0
        metrics['equity_ratio'] = (metrics['equity'] / (metrics['equity'] + metrics['liabilities'])) if (metrics['equity'] + metrics['liabilities']) > 0 else 0.0

        metrics['current_ratio'] = (metrics['assets'] / metrics['liabilities']) if metrics['liabilities'] > 0 else 999.99
        
        metrics['asset_turnover'] = (metrics['revenue'] / metrics['assets']) if metrics['assets'] > 0 else 0.0
        metrics['roa'] = (metrics['profit'] / metrics['assets'] * 100) if metrics['assets'] > 0 else 0.0
        metrics['roe'] = (metrics['profit'] / metrics['equity'] * 100) if metrics['equity'] > 0 else 0.0

        monthly_expenses = metrics['expenses'] / 12 if metrics['expenses'] > 0 else 0.0
        metrics['burn_rate'] = monthly_expenses

        metrics['runway_months'] = (metrics['cash'] / monthly_expenses) if monthly_expenses > 0 else 999.99

        return metrics

    def _identify_key_financial_pages(self, pdf_path, max_pages_to_scan=100, max_tables_to_extract_per_doc=25):
        """
        Identifies pages likely to contain key financial tables (Balance Sheet, P&L, Cash Flow, Notes).
        Prioritizes pages with specific keywords and limits the total number of pages considered.
        Returns a list of 0-indexed page numbers.
        """
        key_pages = set()
        keywords = {
            'balance sheet': ['balance sheet', 'statement of financial position'],
            'income statement': ['income statement', 'statement of profit and loss', 'profit and loss account', 'statement of comprehensive income'],
            'cash flow statement': ['cash flow statement', 'statement of cash flows'],
            'financial notes': ['notes to financial statements', 'summary of significant accounting policies'],
            'audit_report_context': ['auditor\'s report', 'independent auditor\'s report']  
        }
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                num_pdf_pages = len(pdf.pages)
                
          
                pages_to_text_scan = min(num_pdf_pages, max_pages_to_scan)

                for i in range(pages_to_text_scan):
                    page = pdf.pages[i]
                    text = page.extract_text(layout=False)  
                    if text:
                        text_lower = text.lower()
                        
                        found_financial_statement_header = False
                        for k_type, k_list in keywords.items():
                            for k in k_list:
                                if k in text_lower:
                                    logging.debug(f"Page {i+1} identified as potential key page due to keyword: '{k}'")
                                    key_pages.add(i)
                                    if k_type in ['balance sheet', 'income statement', 'cash flow statement']:
                                        found_financial_statement_header = True
                                    break
                            if found_financial_statement_header:
                                break 

                sorted_key_pages = sorted(list(key_pages))
                
            
                if len(sorted_key_pages) < 5 and num_pdf_pages > 0:
                    logging.info(f"Few key pages identified ({len(sorted_key_pages)}). Expanding table scan to first {min(num_pdf_pages, 10)} pages.")
                    additional_pages = set(range(min(num_pdf_pages, 10)))
                    sorted_key_pages = sorted(list(set(sorted_key_pages).union(additional_pages)))

                return sorted_key_pages[:max_tables_to_extract_per_doc]
        except Exception as e:
            logging.error(f"Error identifying key financial pages in {pdf_path}: {e}")
          
          
            try:
                with pdfplumber.open(pdf_path) as pdf_fallback:
                    return list(range(min(len(pdf_fallback.pages), max_tables_to_extract_per_doc)))
            except Exception:
                return []


    def process_file(self, file_path):
        """Main processing method with optimized metric extraction flow."""
        start_time = time.time()
        file_extension = file_path.rsplit('.', 1)[1].lower()
        original_file_name = os.path.basename(file_path)

        text_content = ""
        tables = []
        extracted_raw_data = {'text': '', 'tables': [], 'images_text': ''}

        try:
            if file_extension == 'pdf':
            
                initial_extracted_pdf_data = self.extract_from_pdf_advanced(file_path)
                text_content = initial_extracted_pdf_data.get('text', '') + initial_extracted_pdf_data.get('images_text', '')
                extracted_raw_data['text'] = initial_extracted_pdf_data.get('text', '')
                extracted_raw_data['images_text'] = initial_extracted_pdf_data.get('images_text', '')
                
             
                logging.info(f"Attempting initial traditional metric extraction from text for {original_file_name}...")
                base_metrics_text_only = self.extract_metrics_traditional(text_content, tables=[]) 
                
                prelim_confidence = self._calculate_confidence_score(base_metrics_text_only, text_content, [])
                logging.info(f"Preliminary confidence score (text-only): {prelim_confidence}")
 
                if prelim_confidence < 70:
                    logging.info(f"Preliminary confidence is low, proceeding with intelligent pdfplumber table extraction for {original_file_name}...")
                    
                    table_extraction_start_time = time.time()
                    try:
                        pages_to_extract_tables_from = self._identify_key_financial_pages(
                            file_path,
                            max_pages_to_scan=100, # Scan up to 100 pages for keywords
                            max_tables_to_extract_per_doc=25 # Extract tables from up to 25 identified important pages
                        )
                        logging.info(f"Identified {len(pages_to_extract_tables_from)} pages for table extraction.")

                        with pdfplumber.open(file_path) as pdf:
                            for page_num_idx in pages_to_extract_tables_from:
                                if page_num_idx >= len(pdf.pages): # Ensure page index is valid
                                    continue 
                                page = pdf.pages[page_num_idx]
                                page_tables = page.extract_tables(table_settings={
                                    "vertical_strategy": "lines",
                                    "horizontal_strategy": "lines",
                                    "snap_tolerance": 3,
                                    "join_tolerance": 3,
                                    "edge_min_length": 3,
                                    "min_words_vertical": 3,
                                    "min_words_horizontal": 1,
                                    "text_tolerance": 1,
                                    "text_strategy": "lines"
                                })
                                for table_data in page_tables:
                                    if table_data and len(table_data) > 1:
                                        try:
                                            is_header_row = all(isinstance(x, str) or (pd.isna(x) if pd else x is None) for x in table_data[0])
                                            if is_header_row:
                                                df_table = pd.DataFrame(table_data[1:], columns=table_data[0])
                                            else:
                                                df_table = pd.DataFrame(table_data)
                                        except Exception as df_err:
                                            logging.debug(f"Error inferring header for pdfplumber table on page {page_num_idx+1}: {df_err}. Using default numeric header.")
                                            df_table = pd.DataFrame(table_data)
                                            
                                        df_table = df_table.dropna(how='all').dropna(axis=1, how='all')
                                        if not df_table.empty:
                                            for col in df_table.columns:
                                                # Apply self.clean_financial_value to each element in the column
                                                df_table[col] = df_table[col].apply(self.clean_financial_value)
                                            tables.append(df_table)
                        logging.info(f"Pdfplumber table extraction complete. Found {len(tables)} tables.")
                    except Exception as e:
                        logging.error(f"Pdfplumber table extraction failed for {original_file_name}: {e}", exc_info=True)
                    
                    logging.info(f"Table extraction for {original_file_name} finished in {time.time() - table_extraction_start_time:.2f}s")
                    
                    if tables:
                        logging.info(f"Re-attempting traditional metric extraction with tables for {original_file_name}...")
                        base_metrics = self.extract_metrics_traditional(text_content, tables=tables)
                    else:
                        base_metrics = base_metrics_text_only
                else:
                    base_metrics = base_metrics_text_only
            
            elif file_extension in ['xlsx', 'xls']:
                extracted_excel_data = self.extract_from_excel_advanced(file_path)
                for sheet_name, df in extracted_excel_data['sheets'].items():
                    text_content += df.to_string(index=False, header=True) + "\n"
                tables = extracted_excel_data['tables']
                # Apply financial cleaning to all extracted Excel tables directly
                for df_table in tables:
                    for col in df_table.columns:
                        df_table[col] = df_table[col].apply(self.clean_financial_value)
                base_metrics = self.extract_metrics_traditional(text_content, tables=tables)
            
            elif file_extension == 'csv':
                extracted_csv_data = self.extract_from_csv(file_path)
                text_content = extracted_csv_data['text']
                tables = extracted_csv_data['tables']
                # Apply financial cleaning to all extracted CSV tables directly
                for df_table in tables:
                    for col in df_table.columns:
                        df_table[col] = df_table[col].apply(self.clean_financial_value)
                base_metrics = self.extract_metrics_traditional(text_content, tables=tables)

            metrics = self._calculate_derived_metrics(base_metrics)
            metrics['confidence_score'] = self._calculate_confidence_score(metrics, text_content, tables)
            processing_time = (time.time() - start_time) * 1000
            metrics['processing_time'] = processing_time
            metrics['file_name'] = original_file_name
            
            logging.info(f"Successfully processed {original_file_name} in {processing_time:.0f}ms.")
            return metrics, tables, extracted_raw_data

        except Exception as e:
            logging.error(f"Error processing {original_file_name}: {e}", exc_info=True)
            raise ValueError(f"Error processing file {original_file_name}: {str(e)}")
            
    def _calculate_confidence_score(self, metrics, text_content, tables):
        """Calculate a basic confidence score for extracted data based on completeness and presence."""
        score = 0
        
        critical_metrics = ['revenue', 'profit', 'cash', 'assets', 'liabilities', 'equity']
        found_critical_metrics = sum(1 for key in critical_metrics if metrics.get(key) not in [0.0, "Unknown", None, ""])
        
        score += (found_critical_metrics / len(critical_metrics)) * 50 if critical_metrics else 0

        found_keywords = sum(1 for key in self.financial_keywords if metrics.get(key) not in [0.0, "Unknown", None, ""])
        score += (found_keywords / len(self.financial_keywords)) * 30 if self.financial_keywords else 0

        score += min(20, len(tables) * 5)

        return max(0, min(100, int(score)))


class AdvancedAnalysisAgent:
    """Enhanced analysis agent with advanced ML models and forecasting"""

    def __init__(self):
        self.models = {
            'revenue_rf': RandomForestRegressor(n_estimators=100, random_state=42),
            'profit_gbr': GradientBoostingRegressor(n_estimators=100, learning_rate=0.1, max_depth=3, random_state=42),
            'cash_lr': LinearRegression()
        }
        self.scaler = StandardScaler()
        self.industry_benchmarks = {
            'profitability': {'good': 15, 'average': 8, 'poor': 2},
            'current_ratio': {'good': 2.0, 'average': 1.5, 'poor': 1.0},
            'debt_to_equity': {'good': 0.5, 'average': 1.0, 'poor': 2.0},
            'roa': {'good': 10, 'average': 5, 'poor': 1},
            'roe': {'good': 15, 'average': 10, 'poor': 5},
            'revenue_growth': {'good': 20, 'average': 10, 'poor': 0},
            'cash_ratio': {'good': 0.5, 'average': 0.2, 'poor': 0.1},
        }

    def detect_company_relationships(self, reports):
        """Detect if reports belong to the same company based on name similarity and financial year."""
        relationships = {}
        
        cleaned_report_names_with_indices = []
        for i, report in enumerate(reports):
            if 'metrics' in report and 'company_name' in report['metrics']:
                name = report['metrics'].get('company_name', '').lower()
                if name and name != "unknown":
                    cleaned_name = re.sub(r'\b(ltd|limited|inc|corp|corporation|group|plc|pvt|india)\b', '', name).strip()
                    cleaned_report_names_with_indices.append((cleaned_name, i))
            else:
                logging.debug(f"Skipping report at index {i} due to missing 'metrics' or 'company_name' key.")
                continue

        for i, (name1_cleaned, original_idx1) in enumerate(cleaned_report_names_with_indices):
            if name1_cleaned not in relationships:
                relationships[name1_cleaned] = set()
            relationships[name1_cleaned].add(original_idx1)

            for j, (name2_cleaned, original_idx2) in enumerate(cleaned_report_names_with_indices):
                if original_idx1 >= original_idx2: continue
                
                distance = Levenshtein.distance(name1_cleaned, name2_cleaned)
                max_len = max(len(name1_cleaned), len(name2_cleaned))
                similarity_score = 1 - (distance / max_len) if max_len > 0 else 0

                if similarity_score > 0.7:
                    year1 = reports[original_idx1]['metrics'].get('financial_year')
                    year2 = reports[original_idx2]['metrics'].get('financial_year')
                    if year1 and year2 and year1 != year2:
                        relationships[name1_cleaned].add(original_idx1)
                        relationships[name1_cleaned].add(original_idx2)
                        
                        if name2_cleaned not in relationships:
                            relationships[name2_cleaned] = set()
                        relationships[name2_cleaned].add(original_idx1)
                        relationships[name2_cleaned].add(original_idx2)

        final_relationships = {}
        for company_key, indices_set in relationships.items():
            merged_into_existing = False
            for final_key, final_indices_set in final_relationships.items():
                if not final_indices_set.isdisjoint(indices_set):
                    final_relationships[final_key].update(indices_set)
                    merged_into_existing = True
                    break
            
            if not merged_into_existing:
                first_idx = next(iter(indices_set))
                representative_name = reports[first_idx]['metrics'].get('company_name', company_key.title())
                final_relationships[representative_name] = indices_set
        
        return {k: sorted(list(v)) for k, v in final_relationships.items()}

    def analyze_comprehensive_health(self, metrics):
        """Comprehensive financial health analysis."""
        analysis = {
            'health_score': 0,
            'risk_level': 'low',
            'risks': [],
            'opportunities': [],
            'kpis': {},
            'benchmarks': self.industry_benchmarks
        }

        analysis['kpis'] = self._calculate_comprehensive_kpis(metrics)
        
        score_components = self._calculate_health_components(metrics)
        analysis['health_score'] = self._weighted_health_score(score_components)
        
        analysis['risk_level'] = self._determine_risk_level(analysis['health_score'], metrics)
        
        self._identify_risks(metrics, analysis)
        self._identify_opportunities(metrics, analysis)
        
        return analysis

    def _calculate_health_components(self, metrics):
        """Calculate individual health components with score 0-100."""
        components = {}

        prof = metrics.get('profitability', 0)
        if prof >= 20: components['profitability'] = 100
        elif prof >= 15: components['profitability'] = 85
        elif prof >= 10: components['profitability'] = 70
        elif prof >= 5: components['profitability'] = 55
        elif prof >= 0: components['profitability'] = 40
        else: components['profitability'] = 10

        runway = metrics.get('runway_months', 0)
        current_ratio = metrics.get('current_ratio', 0)
        liquidity_score = 0
        if runway >= 24: liquidity_score += 50
        elif runway >= 12: liquidity_score += 35
        elif runway >= 6: liquidity_score += 20
        else: liquidity_score += 5
        
        if current_ratio >= 2.0: liquidity_score += 50
        elif current_ratio >= 1.5: liquidity_score += 35
        elif current_ratio >= 1.0: liquidity_score += 20
        else: liquidity_score += 5
        components['liquidity'] = min(100, liquidity_score)

        debt_eq = metrics.get('debt_to_equity', 0)
        if debt_eq <= 0.3: components['leverage'] = 100
        elif debt_eq <= 0.5: components['leverage'] = 85
        elif debt_eq <= 1.0: components['leverage'] = 70
        elif debt_eq <= 1.5: components['leverage'] = 55
        elif debt_eq <= 2.0: components['leverage'] = 40
        else: components['leverage'] = 10
        
        return components

    def _weighted_health_score(self, components):
        """Calculate weighted health score from components."""
        weights = {'profitability': 0.40, 'liquidity': 0.35, 'leverage': 0.25}
        weighted_score = sum(components.get(key, 0) * weights[key] for key in weights)
        return max(0, min(100, int(weighted_score)))

    def _determine_risk_level(self, health_score, metrics):
        """Determine risk level with more granular assessment."""
        if health_score >= 85:
            return 'low'
        elif health_score >= 70:
            return 'medium'
        elif health_score >= 50:
            return 'high'
        else:
            return 'critical'

    def _identify_risks(self, metrics, analysis):
        """Identify comprehensive financial risks with actionable outputs."""
        risks = []

        runway = metrics.get('runway_months', 0)
        if runway < 3 and metrics.get('cash', 0) <= 0:
            risks.append({
                'level': 'critical',
                'type': 'liquidity',
                'message': f'Imminent Cash Crisis (Negative Cash / {runway:.1f} months runway)',
                'impact': 'Company is highly likely to cease operations or face severe distress within 3 months, unable to meet short-term obligations.',
                'recommendation': 'Secure emergency bridge funding immediately, initiate drastic cost-cutting across all departments, and aggressively accelerate receivables collection. Consider asset liquidation.',
                'timeline': 'Immediate (within days to 2 weeks)',
                'kpi_affected': ['Runway', 'Cash Reserves', 'Current Ratio']
            })
        elif runway < 3:
            risks.append({
                'level': 'critical',
                'type': 'liquidity',
                'message': f'Imminent Cash Crisis ({runway:.1f} months runway)',
                'impact': 'Company may cease operations or face severe distress within 3 months, unable to meet short-term obligations.',
                'recommendation': 'Secure emergency bridge funding immediately, initiate drastic cost-cutting, and accelerate receivables.',
                'timeline': 'Immediate (within days)',
                'kpi_affected': ['Runway', 'Cash Reserves']
            })
        elif runway < 6:
            risks.append({
                'level': 'high',
                'type': 'liquidity',
                'message': f'Critical Cash Shortage ({runway:.1f} months runway)',
                'impact': 'Company may struggle to meet obligations within 6 months, limiting strategic options and increasing borrowing costs.',
                'recommendation': 'Implement emergency cost reduction, explore short-term financing (e.g., line of credit), and optimize working capital. Renegotiate payment terms with suppliers.',
                'timeline': 'Short-term (1-3 months)',
                'kpi_affected': ['Runway', 'Cash Reserves', 'Current Ratio']
            })
        elif runway < 12:
            risks.append({
                'level': 'medium',
                'type': 'liquidity',
                'message': f'Limited Cash Runway ({runway:.1f} months)',
                'impact': 'Insufficient buffer for unexpected challenges or investment opportunities, restricting growth potential and increasing reliance on external financing.',
                'recommendation': 'Improve cash flow forecasting, negotiate better payment terms with suppliers/customers, and explore non-dilutive funding. Focus on extending payment terms.',
                'timeline': 'Medium-term (3-6 months)',
                'kpi_affected': ['Runway', 'Cash Reserves']
            })

        current_ratio = metrics.get('current_ratio', 0)
        if current_ratio < self.industry_benchmarks['current_ratio']['poor']:
            risks.append({
                'level': 'high',
                'type': 'liquidity',
                'message': f'Poor Liquidity (Current Ratio: {current_ratio:.2f})',
                'impact': 'May struggle to meet short-term liabilities with current assets, risking operational disruption, default, and reputational damage.',
                'recommendation': 'Improve working capital management, liquidate non-essential current assets, or secure short-term credit facilities. Review inventory and accounts receivable turnover.',
                'timeline': 'Short-term (1-3 months)',
                'kpi_affected': ['Current Ratio']
            })
        elif current_ratio < self.industry_benchmarks['current_ratio']['average']:
            risks.append({
                'level': 'medium',
                'type': 'liquidity',
                'message': f'Tight Liquidity (Current Ratio: {current_ratio:.2f})',
                'impact': 'Limited financial flexibility, making the company vulnerable to unexpected expenses or revenue dips. Growth initiatives may be constrained.',
                'recommendation': 'Monitor cash flow closely, optimize inventory, and potentially extend payables while managing receivables efficiently. Explore dynamic discounting for early payments.',
                'timeline': 'Short to Medium-term (3-6 months)',
                'kpi_affected': ['Current Ratio']
            })

        profitability = metrics.get('profitability', 0)
        if profitability < self.industry_benchmarks['profitability']['poor'] and profitability < 0:
            risks.append({
                'level': 'critical',
                'type': 'profitability',
                'message': f'Severe Losses ({profitability:.1f}% margin)',
                'impact': 'Unsustainable business model, rapid depletion of equity, and high risk of insolvency if not addressed immediately. Investors will lose confidence.',
                'recommendation': 'Conduct a comprehensive review of pricing, product strategy, and operating costs. Identify and eliminate unprofitable segments. Implement aggressive cost control measures.',
                'timeline': 'Immediate',
                'kpi_affected': ['Profitability', 'Net Profit']
            })
        elif profitability < self.industry_benchmarks['profitability']['poor']:
            risks.append({
                'level': 'high',
                'type': 'profitability',
                'message': f'Low Profitability ({profitability:.1f}% margin)',
                'impact': 'Limited capacity for reinvestment, debt repayment, and dividend distribution, hindering long-term growth and market competitiveness.',
                'recommendation': 'Focus on increasing gross margins through better procurement or pricing, and identify areas for operational efficiency. Analyze competitor pricing and product value.',
                'timeline': 'Short-term (3-6 months)',
                'kpi_affected': ['Profitability', 'Net Profit']
            })

        debt_to_equity = metrics.get('debt_to_equity', 0)
        if debt_to_equity > self.industry_benchmarks['debt_to_equity']['poor'] * 1.5:
            risks.append({
                'level': 'critical',
                'type': 'leverage',
                'message': f'Excessive Leverage (D/E: {debt_to_equity:.2f})',
                'impact': 'High financial risk, potential for covenant breaches, making new financing difficult and expensive. Risk of bankruptcy is significantly elevated.',
                'recommendation': 'Urgent debt restructuring, equity infusion, or sale of non-core assets to reduce debt. Renegotiate terms with lenders.',
                'timeline': 'Immediate',
                'kpi_affected': ['Debt-to-Equity']
            })
        elif debt_to_equity > self.industry_benchmarks['debt_to_equity']['poor']:
            risks.append({
                'level': 'high',
                'type': 'leverage',
                'message': f'High Leverage (D/E: {debt_to_equity:.2f})',
                'impact': 'Limited financial flexibility, increased interest burden, and vulnerability to interest rate fluctuations. May impact credit rating and future borrowing capacity.',
                'recommendation': 'Develop a clear debt reduction strategy, prioritize cash flow towards debt repayment, and consider refinancing. Explore options for converting short-term to long-term debt.',
                'timeline': 'Short to Medium-term (6-12 months)',
                'kpi_affected': ['Debt-to-Equity']
            })
        
        if metrics.get('profit', 0) < 0 and metrics.get('revenue', 0) < 0:
             risks.append({
                'level': 'critical',
                'type': 'operational',
                'message': 'Negative Revenue and Profit',
                'impact': 'Indicates severe operational failure and potential data inaccuracy or business collapse.',
                'recommendation': 'Immediate forensic accounting review and strategic pivot or cessation of operations.',
                'timeline': 'Immediate',
                'kpi_affected': ['Revenue', 'Profit']
            })
        elif metrics.get('profit', 0) < 0:
            risks.append({
                'level': 'high',
                'type': 'operational',
                'message': 'Sustained Operating Losses',
                'impact': 'The core business is not generating enough profit to cover costs, eroding equity and threatening long-term viability.',
                'recommendation': 'Re-evaluate business model, operational efficiency, and market positioning. Consider product portfolio optimization and competitive analysis.',
                'timeline': 'Short to Medium-term (3-9 months)',
                'kpi_affected': ['Profit', 'Profitability']
            })

        if not risks:
            risks.append({
                'level': 'low',
                'type': 'overall',
                'message': 'No significant risks detected.',
                'impact': 'The company maintains a stable financial position with adequate safeguards, allowing for strategic planning.',
                'recommendation': 'Continue to monitor market conditions and maintain prudent financial management. Focus on sustainable growth.',
                'timeline': 'Ongoing',
                'kpi_affected': ['All KPIs']
            })

        analysis['risks'] = risks

    def _identify_opportunities(self, metrics, analysis):
        """Identify financial opportunities based on strong metrics."""
        opportunities = []

        if metrics.get('profitability', 0) > self.industry_benchmarks['profitability']['good']:
            opportunities.append({
                'type': 'growth',
                'message': f'Strong Profitability ({metrics["profitability"]:.1f}% margin)',
                'impact': 'Enables leveraging profits for strategic investments, market expansion, or product development without excessive external debt.',
                'recommendation': 'Consider strategic investments in R&D, market expansion, or potential acquisitions to capitalize on competitive advantage. Explore new revenue streams.',
                'timeline': 'Medium-term (6-18 months)'
            })
        
        if metrics.get('runway_months', 0) > 18 and metrics.get('cash', 0) > 0:
            opportunities.append({
                'type': 'investment',
                'message': f'Excellent Cash Reserves (₹{metrics["cash"]:.2f} Cr, {metrics["runway_months"]:.1f} months runway)',
                'impact': 'Provides strong financial resilience, flexibility for opportunistic investments, potential for share buybacks or dividends, and a strong negotiation position.',
                'recommendation': 'Evaluate potential mergers & acquisitions, accelerate debt repayment, or initiate a share repurchase program. Explore optimizing idle cash through short-term, low-risk investments.',
                'timeline': 'Long-term (1-3 years)'
            })
        
        if metrics.get('debt_to_equity', 0) < self.industry_benchmarks['debt_to_equity']['good']:
            opportunities.append({
                'type': 'financing',
                'message': f'Low Leverage (Debt-to-Equity: {metrics["debt_to_equity"]:.2f})',
                'impact': 'Offers significant capacity to take on new debt for strategic initiatives at favorable interest rates, if needed, without jeopardizing financial stability.',
                'recommendation': 'While current leverage is good, this provides flexibility to fund future growth through debt if equity is too expensive or unavailable. Maintain a balanced capital structure.',
                'timeline': 'Long-term (1-3 years)'
            })

        if not opportunities and metrics.get('health_score', 0) >= 70:
            opportunities.append({
                'type': 'general',
                'message': 'Solid Financial Foundation',
                'impact': 'A strong financial position provides stability and a basis for future growth, making the company attractive to investors and partners.',
                'recommendation': 'Continue to innovate and explore new market segments. Invest in customer retention and operational excellence to sustain growth.',
                'timeline': 'Ongoing'
            })

        analysis['opportunities'] = opportunities

    def _calculate_comprehensive_kpis(self, metrics):
        """Calculate a comprehensive set of KPIs."""
        kpis = {}
        
        kpis['gross_margin'] = metrics.get('gross_margin', 0.0)
        kpis['net_profit_margin'] = metrics.get('profitability', 0.0)
        kpis['ebitda'] = metrics.get('profit',0) + metrics.get('interest_expense',0) + metrics.get('tax_expense',0) + metrics.get('depreciation_amortization',0) # Placeholder for more complex calculation
        kpis['ebitda_margin'] = (kpis['ebitda'] / metrics.get('revenue',1)) * 100 if metrics.get('revenue',0)>0 else 0.0
        
        kpis['current_ratio'] = metrics.get('current_ratio', 0.0)
        kpis['quick_ratio'] = ((metrics.get('assets',0) * 0.8) / metrics.get('liabilities',1)) if metrics.get('liabilities',0)>0 else 999.99
        kpis['cash_ratio'] = (metrics.get('cash',0) / metrics.get('liabilities',1)) if metrics.get('liabilities',0)>0 else 999.99
        kpis['runway_months'] = metrics.get('runway_months', 0.0)

        kpis['asset_turnover'] = metrics.get('asset_turnover', 0.0)
        kpis['inventory_turnover'] = np.random.uniform(4, 12)
        kpis['receivables_turnover'] = np.random.uniform(6, 12)
        
        kpis['debt_to_equity'] = metrics.get('debt_to_equity', 0.0)
        kpis['debt_to_assets'] = (metrics.get('liabilities',0) / metrics.get('assets',1)) if metrics.get('assets',0)>0 else 0.0
        kpis['equity_multiplier'] = (metrics.get('assets',0) / metrics.get('equity',1)) if metrics.get('equity',0)>0 else 0.0

        kpis['roa'] = metrics.get('roa', 0.0)
        kpis['roe'] = metrics.get('roe', 0.0)

        kpis['revenue_growth'] = np.random.uniform(5, 25) if metrics.get('revenue',0)>0 else 0.0
        kpis['profit_growth'] = np.random.uniform(5, 30) if metrics.get('profit',0)>0 else 0.0

        return kpis

    def predict_advanced_trends(self, metrics, forecast_years=5):
        """Advanced trend prediction using Prophet model or fallback sklearn models for 3-5 years."""
        predictions = {}
        total_months = forecast_years * 12

        base_date = datetime.now() - timedelta(days=365 * 3)
        dates = pd.to_datetime([base_date + timedelta(days=30 * i) for i in range(1, 37)])

        initial_revenue = metrics.get('revenue', 1000.0) / (1 + 0.15/12)**35
        initial_profit = metrics.get('profit', 100.0) / (1 + 0.20/12)**35
        initial_cash = metrics.get('cash', 500.0) / (1 + 0.10/12)**35

        historical_data = pd.DataFrame({
            'ds': dates,
            'revenue': [initial_revenue * (1 + 0.15/12)**i + np.random.normal(0, initial_revenue*0.01) for i in range(36)],
            'profit': [initial_profit * (1 + 0.20/12)**i + np.random.normal(0, initial_profit*0.02) for i in range(36)],
            'cash': [initial_cash * (1 + 0.10/12)**i + np.random.normal(0, initial_cash*0.01) for i in range(36)],
        })

        historical_data.loc[historical_data.index[-1], 'revenue'] = metrics.get('revenue', 1000.0)
        historical_data.loc[historical_data.index[-1], 'profit'] = metrics.get('profit', 100.0)
        historical_data.loc[historical_data.index[-1], 'cash'] = metrics.get('cash', 500.0)

        historical_data['revenue'] = historical_data['revenue'].apply(lambda x: max(0, x))
        historical_data['cash'] = historical_data['cash'].apply(lambda x: max(0, x))


        scenarios = {
            'conservative': {'revenue_factor': 0.9, 'profit_factor': 0.8, 'cash_impact': 0.9},
            'realistic': {'revenue_factor': 1.0, 'profit_factor': 1.0, 'cash_impact': 1.0},
            'optimistic': {'revenue_factor': 1.1, 'profit_factor': 1.2, 'cash_impact': 1.1}
        }

        prophet_is_available_for_this_run = PROPHET_AVAILABLE 

        for metric_name in ['revenue', 'profit', 'cash']:
            hist_series_df = historical_data[['ds', metric_name]].rename(columns={metric_name: 'y'})
            
            hist_series_df['y'] = pd.to_numeric(hist_series_df['y'], errors='coerce').fillna(0.0)

            forecast_values_base = np.array([])

            if prophet_is_available_for_this_run:
                try:
                    model = Prophet(
                        growth='linear',
                        yearly_seasonality=True,
                        weekly_seasonality=False,
                        daily_seasonality=False,
                        seasonality_mode='multiplicative',
                        changepoint_prior_scale=0.05
                    )
                    model.fit(hist_series_df)
                    
                    future = model.make_future_dataframe(periods=total_months, freq='M')
                    forecast = model.predict(future)
                    
                    forecast_values_base = forecast['yhat'].iloc[-total_months:].values
                    logging.info(f"Prophet forecasting successful for {metric_name}.")
                except Exception as e:
                    logging.warning(f"Prophet forecasting failed for {metric_name}: {e}. Falling back to sklearn models.")
                    prophet_is_available_for_this_run = False
            
            if not prophet_is_available_for_this_run:
                logging.info(f"Using sklearn fallback for {metric_name} prediction.")
                X_hist = np.arange(len(hist_series_df)).reshape(-1, 1)
                y_hist = hist_series_df['y'].values
                
                if metric_name == 'revenue':
                    self.models['revenue_rf'].fit(X_hist, y_hist)
                    base_sklearn_model = self.models['revenue_rf']
                elif metric_name == 'profit':
                    self.models['profit_gbr'].fit(X_hist, y_hist)
                    base_sklearn_model = self.models['profit_gbr']
                else:
                    self.models['cash_lr'].fit(X_hist, y_hist)
                    base_sklearn_model = self.models['cash_lr']

                future_months_indices = np.arange(len(hist_series_df), len(hist_series_df) + total_months).reshape(-1, 1)
                forecast_values_base = base_sklearn_model.predict(future_months_indices)

            for scenario_name, factors in scenarios.items():
                factor_key = f"{metric_name}_factor"
                impact_factor = factors.get(factor_key, 1.0) if metric_name != 'cash' else factors.get('cash_impact', 1.0)
                
                forecast_scenario = forecast_values_base * impact_factor
                
                if metric_name in ['revenue', 'cash']:
                    forecast_scenario = np.maximum(0, forecast_scenario)
                
                predictions[f'{metric_name}_{scenario_name}'] = {
                    'values': forecast_scenario.tolist(),
                    'trend': 'increasing' if forecast_scenario[-1] > forecast_scenario[0] else 'decreasing'
                }

        predictions['revenue'] = predictions['revenue_realistic']
        predictions['profit'] = predictions['profit_realistic']
        predictions['cash'] = predictions['cash_realistic']
        
        return predictions

    def compare_reports(self, reports):
        """Compare multiple reports from the same company (different periods)."""
        if len(reports) < 2:
            return {}

        comparison = {
            'metrics_over_time': {},
            'growth_analysis': {},
            'trend_summary': {}
        }

        metrics_by_year = {}
        for report in reports:
            if 'metrics' in report:
                year_str = report['metrics'].get('financial_year', 'Unknown').split('-')[0]
                if year_str != 'Unknown' and year_str.isdigit():
                    metrics_by_year[int(year_str)] = {
                        'report_id': report.get('id', None),
                        'file_name': report.get('file_name', 'N/A'),
                        'revenue': report['metrics'].get('revenue', 0),
                        'profit': report['metrics'].get('profit', 0),
                        'cash': report['metrics'].get('cash', 0),
                        'assets': report['metrics'].get('assets', 0),
                        'profitability': report['metrics'].get('profitability', 0),
                        'health_score': report['analysis'].get('health_score', 0),
                        'risk_level': report['analysis'].get('risk_level', 'unknown'),
                        'runway_months': report['metrics'].get('runway_months', 0)
                    }

        sorted_years = sorted(metrics_by_year.keys())
        comparison['metrics_over_time'] = {str(year): metrics_by_year[year] for year in sorted_years}

        growth_metrics = ['revenue', 'profit', 'cash', 'assets']
        for i in range(1, len(sorted_years)):
            prev_year_data = metrics_by_year[sorted_years[i-1]]
            curr_year_data = metrics_by_year[sorted_years[i]]
            period_label = f"{sorted_years[i-1]} to {sorted_years[i]}"

            for metric in growth_metrics:
                prev_val = prev_year_data.get(metric, 0)
                curr_val = curr_year_data.get(metric, 0)
                
                if prev_val > 0:
                    growth_rate = ((curr_val - prev_val) / prev_val) * 100
                elif curr_val > 0:
                    growth_rate = 100.0
                else:
                    growth_rate = 0.0

                if metric not in comparison['growth_analysis']:
                    comparison['growth_analysis'][metric] = []
                comparison['growth_analysis'][metric].append({
                    'period': period_label,
                    'growth_rate': round(growth_rate, 2),
                    'from_value': round(prev_val, 2),
                    'to_value': round(curr_val, 2)
                })
        
        for metric in growth_metrics:
            if metric in comparison['growth_analysis'] and comparison['growth_analysis'][metric]:
                avg_growth = np.mean([g['growth_rate'] for g in comparison['growth_analysis'][metric]])
                comparison['trend_summary'][metric] = f"Average annual growth of {avg_growth:.2f}%"
                if avg_growth > 0:
                    comparison['trend_summary'][metric] += " (Positive trend)"
                elif avg_growth < 0:
                    comparison['trend_summary'][metric] += " (Negative trend)"
                else:
                    comparison['trend_summary'][metric] += " (Stable)"

        logging.info(f"Comparison of {len(reports)} reports complete.")
        return comparison

class AdvancedAdvisorAgent:
    """Enhanced advisor agent with Ollama-powered or traditional rule-based insights and user interaction."""

    def __init__(self):
        self.chat_history = []

    def generate_executive_report(self, metrics, analysis, predictions, company_comparison=None):
        """Generate comprehensive executive report using Ollama or traditional templating."""
        if USE_OLLAMA:
            return self._generate_ollama_report(metrics, analysis, predictions, company_comparison)
        else:
            return self._generate_traditional_report(metrics, analysis, predictions, company_comparison)

    def _generate_ollama_report(self, metrics, analysis, predictions, company_comparison=None):
        """Generate an executive report using an Ollama LLM."""
        company_name = metrics.get('company_name', 'Your Company')
        financial_year = metrics.get('financial_year', 'Current Period')

        report_context  = f"""
You are an AI CFO Assistant for {company_name}.

Latest Financial Metrics:
- Revenue: ₹{metrics.get('revenue', 0.0):.2f} Cr
- Profit: ₹{metrics.get('profit', 0.0):.2f} Cr
- Health Score: {analysis.get('health_score', 0):.0f}/100
- Risk Level: {analysis.get('risk_level', 'unknown').upper()}

⚖ Role & Rules:

don't use these strictly :unknown,Unknown,Unknown's,unknown's keyword use user instead
1. Only respond to user questions strictly related to financial report metrics such as:
   Revenue, Profit, Health Score, Risk Level, Runway, Cash Reserves, Burn Rate, and related KPIs.
2. If the user asks anything unrelated (e.g., personal, general knowledge, casual chat), strictly reply:
   "Please ask only about the financial report metrics (Revenue, Profit, Health Score, Risk Level, Runway, Cash Reserves, Burn Rate)."
3. Responses must be short, concise and to the point (1-5 lines maximum), clear, and fast — like a professional CFO summary.
4. For questions about future outlook, trends, or recommendations:
   - Provide a *short forward-looking analysis* (1-2 lines).
   - Add a *CFO-style recommendation* (1-2 lines) focusing on risk management, efficiency, or growth.
 Guidance for Future Recommendations:
- If Revenue is strong but Profit is weak → recommend cost optimization or margin improvement.
- If Profit is strong but Cash is low → recommend better working capital or liquidity management.
- If Health Score is low → recommend risk mitigation, expense control, or fresh capital infusion.
- If Risk Level is HIGH → recommend diversification, hedging, or strategic reserves.

 Previous conversation:
"""
        if analysis['risks']:
            for risk in analysis['risks']:
                report_context += f"- **{risk['level'].upper()} Risk ({risk['type']}):** {risk['message']}.\n  Impact: {risk['impact']}\n  Recommendation: {risk['recommendation']}\n  Timeline: {risk['timeline']}\n  KPIs Affected: {', '.join(risk.get('kpi_affected', ['N/A']))}\n"
        else:
            report_context += "- No significant risks identified.\n"
        
        report_context += """
        **Identified Opportunities:**
        """
        if analysis['opportunities']:
            for opp in analysis['opportunities']:
                report_context += f"- **Opportunity ({opp['type']}):** {opp['message']}.\n  Impact: {opp['impact']}\n  Recommendation: {opp['recommendation']}\n  Timeline: {opp['timeline']}\n"
        else:
            report_context += "- No specific opportunities highlighted at this time.\n"

        report_context += f"""
        **Future Predictions (Realistic Scenario for next 5 years):**
        Projected Revenue Trend: {'increasing' if predictions['revenue']['trend'] == 'increasing' else 'decreasing'}
        Projected Profit Trend: {'increasing' if predictions['profit']['trend'] == 'increasing' else 'decreasing'}
        Projected Cash Trend: {'increasing' if predictions['cash']['trend'] == 'increasing' else 'decreasing'}
        """
        
        if company_comparison:
            report_context += "\n**Multi-Report Comparison Summary (Year-over-Year):**\n"
            for metric, trend_summary in company_comparison.get('trend_summary', {}).items():
                report_context += f"- {metric.replace('_', ' ').title()}: {trend_summary}.\n"
        
        report_context += f"""
        Please generate a full executive report based on this information, structured with the following sections:
        1. **Executive Summary**: A concise overview of financial health, key performance, and outlook.
        2. **Key Financial Highlights**: Presenting the most important financial metrics.
        3. **Detailed Risk Assessment & Mitigation Strategies**: List major risks, their impacts, and specific, actionable recommendations.
        4. **Strategic Opportunities & Growth Initiatives**: Highlight potential areas for growth and investment based on financial strengths.
        5. **Forward-Looking Outlook & Predictions**: Summarize future trends and forecasts.
        6. **Actionable Recommendations Summary**: A bulleted list of 3-5 high-priority actions.
        Ensure the tone is professional, directly addresses the implications of the numbers, and uses clear, concise language.
        """

        try:
            ollama_payload = {
                "model": OLLAMA_MODEL,
                "prompt": report_context,
                "stream": False,
                "options": {
                    "temperature": 0.3,
                    "num_ctx": 4096 # Adjust context window if needed for very long reports
                }
            }
            # Increased timeout to 600 seconds (10 minutes) for report generation
            response = requests.post(f"{OLLAMA_API_BASE_URL}/api/generate", json=ollama_payload, timeout=600)
            response.raise_for_status()
            ollama_output = response.json()['response'].strip()
            return ollama_output
        except requests.exceptions.RequestException as e:
            logging.error(f"Ollama report generation failed: {e}", exc_info=True)
            return self._generate_traditional_report(metrics, analysis, predictions, company_comparison) # Fallback

    def _generate_traditional_report(self, metrics, analysis, predictions, company_comparison=None):
        """Generate traditional template-based report."""
        company_name = metrics.get('company_name', 'Company')
        financial_year = metrics.get('financial_year', 'Current Period')

        report_content = f"""
**EXECUTIVE SUMMARY - {company_name}**
Financial Year: {financial_year}

{company_name} demonstrates a **{analysis.get('risk_level', 'unknown').upper()}** risk profile with an overall Financial Health Score of **{analysis.get('health_score', 0):.0f}/100**.
The company reported a revenue of **₹{metrics.get('revenue', 0):.2f} Crores** and maintained a profitability margin of **{metrics.get('profitability', 0):.1f}%**.
Cash reserves stand at **₹{metrics.get('cash', 0):.2f} Crores**, providing an estimated **{metrics.get('runway_months', 0):.1f} months** of runway.

---

**KEY FINANCIAL HIGHLIGHTS**

*   **Revenue**: ₹{metrics.get('revenue', 0):.2f} Crores
*   **Net Profit**: ₹{metrics.get('profit', 0):.2f} Crores ({metrics.get('profitability', 0):.1f}% margin)
*   **Cash Position**: ₹{metrics.get('cash', 0):.2f} Crores
*   **Cash Runway**: {metrics.get('runway_months', 0):.1f} months (at current burn rate of ₹{metrics.get('burn_rate', 0):.2f} Cr/month)
*   **Current Ratio**: {metrics.get('current_ratio', 0):.2f}
*   **Debt-to-Equity**: {metrics.get('debt_to_equity', 0):.2f}
*   **Return on Assets (ROA)**: {metrics.get('roa', 0):.1f}%
*   **Return on Equity (ROE)**: {metrics.get('roe', 0):.1f}%

---

**DETAILED RISK ASSESSMENT & MITIGATION STRATEGIES**
"""
        if analysis['risks']:
            for risk in analysis['risks']:
                report_content += f"""
*   **{risk['level'].upper()} Risk ({risk['type'].title()})**: {risk['message']}
    *   **Why it matters**: {risk['impact']}
    *   **Recommended next step**: {risk['recommendation']}
    *   **Timeline**: {risk['timeline']}
    *   **KPIs Affected**: {', '.join(risk.get('kpi_affected', ['N/A']))}
"""
        else:
            report_content += "\n*   No significant risks identified in the current analysis. Proactive monitoring is recommended.\n"

        report_content += """
---

**STRATEGIC OPPORTUNITIES & GROWTH INITIATIVES**
"""
        if analysis['opportunities']:
            for opp in analysis['opportunities']:
                report_content += f"""
*   **Opportunity ({opp['type'].title()}):** {opp['message']}
    *   **Impact**: {opp['impact']}
    *   **Recommended next step**: {opp['recommendation']}
    *   **Timeline**: {opp['timeline']}
"""
        else:
            report_content += "\n*   No explicit opportunities highlighted at this time.\n"

        report_content += f"""
---

**FORWARD-LOOKING OUTLOOK & PREDICTIONS (Realistic Scenario for next 5 years)**

*   **Revenue Trend**: Expected to be **{predictions['revenue']['trend']}** over the next 5 years.
*   **Profit Trend**: Expected to be **{predictions['profit']['trend']}** over the next 5 years.
*   **Cash Flow Trend**: Expected to be **{predictions['cash']['trend']}** over the next 5 years.

The financial trajectory indicates {'positive growth' if predictions['revenue']['trend'] == 'increasing' else 'potential challenges'} requiring {'strategic investments' if predictions['revenue']['trend'] == 'increasing' else 'careful management'}.

---

**ACTIONABLE RECOMMENDATIONS SUMMARY**

*   **Cash Flow Management**: {'Prioritize extending cash runway to ensure long-term stability.' if metrics.get('runway_months', 0) < 12 else 'Maintain robust cash flow management to capitalize on opportunities.'}
*   **Profitability Enhancement**: {'Focus on operational efficiencies and revenue growth initiatives to improve margin.' if metrics.get('profitability', 0) < 10 else 'Sustain and optimize current profitability levels.'}
*   **Risk Mitigation**: {'Implement immediate action plans for critical and high-level risks identified.' if any(r['level'] in ['critical', 'high'] for r in analysis['risks']) else 'Continuously monitor key financial indicators for emerging risks.'}
*   **Strategic Growth**: {'Explore strategic investments in new markets or product lines given strong financial health.' if analysis.get('health_score', 0) >= 70 else 'Consolidate current market position before aggressive expansion.'}
"""
        if company_comparison and company_comparison.get('growth_analysis', {}).get('revenue'):
             avg_rev_growth = np.mean([g['growth_rate'] for g in company_comparison['growth_analysis']['revenue']])
             report_content += f"""
*   **Historical Performance**: Review past performance, particularly {company_comparison['trend_summary'].get('revenue', 'revenue trends')}, to inform future strategy.
"""

        return report_content

    def chat_response(self, query, metrics, analysis, chat_history):
        """Generate contextual chat response using Ollama or keyword-based logic."""
        if USE_OLLAMA:
            return self._ollama_chat_response(query, metrics, analysis, chat_history)
        else:
            return self._traditional_chat_response(query, metrics, analysis, predictions=analysis['predictions'])

 





    def _ollama_chat_response(self, query, metrics, analysis, chat_history):
        """AI-powered chat response using Ollama, modified to be a generator for streaming."""
        company_name = metrics.get('company_name', 'the company')
        context_summary = f"""You are an AI CFO Assistant for {company_name}.
        Here are the latest financial metrics:
        - Revenue: ₹{metrics.get('revenue', 0.0):.2f} Cr
        - Profit: ₹{metrics.get('profit', 0.0):.2f} Cr
        - Health Score: {analysis.get('health_score', 0):.0f}/100
        - Risk Level: {analysis.get('risk_level', 'unknown').upper()}
        Previous conversation:
        """
        for msg in chat_history[-6:]:
            context_summary += f"{msg['role'].capitalize()}: {msg['content']}\n"
        context_summary += f"\nUser: {query}\n\nAI CFO Assistant: "

        ollama_payload = {
            "model": OLLAMA_MODEL,
            "prompt": context_summary,
            "stream": True  
        }
        
        try:
   
            response = requests.post(
                f"{OLLAMA_API_BASE_URL}/api/generate", 
                json=ollama_payload, 
                stream=True, 
                timeout=60
            )
            response.raise_for_status()
            for line in response.iter_lines():
                if line:
                    try:
                        chunk = json.loads(line)
                        if not chunk.get('done'):  
                            yield chunk.get('response', '')
                    except json.JSONDecodeError:
                        logging.warning(f"Could not decode JSON line from Ollama stream: {line}")
                        continue
            
  

        except requests.exceptions.RequestException as e:
            logging.error(f"Ollama chat streaming failed: {e}", exc_info=True)
            yield "Sorry, I encountered an error connecting to the AI model. Please check the connection and try again."

    def _traditional_chat_response(self, query, metrics, analysis, predictions):
        """Traditional keyword-based chat response."""
        query_lower = query.lower()
        company_name = metrics.get('company_name', 'Your Company')

        if 'hello' in query_lower or 'hi' in query_lower:
            return f"Hello! I'm your AI CFO Assistant for {company_name}. How can I help you today regarding your financial analysis?"

        elif 'health' in query_lower or 'score' in query_lower:
            health_score = analysis.get('health_score', 0)
            risk_level = analysis.get('risk_level', 'unknown')
            return f"The financial health score for {company_name} is {health_score:.0f}/100, indicating a **{risk_level.upper()}** risk level. This suggests {'excellent performance' if health_score >= 80 else 'a stable position' if health_score >= 60 else 'areas needing attention' if health_score >= 40 else 'a critical situation'}."
        
        elif 'revenue' in query_lower or 'sales' in query_lower or 'income' in query_lower:
            revenue = metrics.get('revenue', 0)
            profitability = metrics.get('profitability', 0)
            return f"{company_name}'s current revenue is ₹{revenue:.2f} Cr with a profitability margin of {profitability:.1f}%. {'This indicates strong top-line performance and efficiency.' if profitability > 10 else 'There is room for improvement in revenue generation or cost management to boost profitability.'}"
        
        elif 'cash' in query_lower or 'runway' in query_lower or 'liquidity' in query_lower:
            cash = metrics.get('cash', 0)
            runway = metrics.get('runway_months', 0)
            burn_rate = metrics.get('burn_rate', 0)
            return f"{company_name} has cash reserves of ₹{cash:.2f} Cr. With a monthly burn rate of ₹{burn_rate:.2f} Cr, the cash runway is approximately {runway:.1f} months. {'This provides a strong financial cushion.' if runway > 12 else 'Consider strategies to extend the runway, such as cost optimization or seeking additional funding.'}"
        
        elif 'profit' in query_lower or 'profitability' in query_lower or 'margin' in query_lower:
            profit = metrics.get('profit', 0)
            profitability = metrics.get('profitability', 0)
            return f"Net profit for {company_name} is ₹{profit:.2f} Cr, translating to a margin of {profitability:.1f}%. {'This is excellent profitability, indicating efficient operations.' if profitability > 15 else 'Good profitability, but continuous efforts are needed to sustain or improve margins.' if profitability > 5 else 'Profitability needs improvement. Review cost structure and pricing strategies.'}"
        
        elif 'debt' in query_lower or 'leverage' in query_lower or 'liabilities' in query_lower:
            debt_to_equity = metrics.get('debt_to_equity', 0)
            return f"The Debt-to-Equity ratio for {company_name} is {debt_to_equity:.2f}. {'This indicates low leverage and a financially conservative approach, offering flexibility.' if debt_to_equity < 0.5 else 'Moderate leverage, which is generally acceptable. Monitor it closely.' if debt_to_equity < 1.0 else 'High leverage. It is crucial to monitor debt levels closely and develop a reduction strategy.'}."
        
        elif 'risk' in query_lower or 'risks' in query_lower or 'concerns' in query_lower:
            risks = analysis.get('risks', [])
            if risks:
                top_risk = risks[0]
                return f"One of the key risks identified for {company_name} is: **{top_risk['message']}**. Its impact is '{top_risk['impact']}'. The recommended action is: '{top_risk['recommendation']}'. "
            else:
                return f"No significant risks detected in {company_name}'s current financial analysis. Your financial position appears stable."
        
        elif 'recommendation' in query_lower or 'advice' in query_lower or 'next step' in query_lower or 'suggest' in query_lower:
            health_score = analysis.get('health_score', 0)
            if health_score >= 80:
                return f"Given {company_name}'s strong financial position, consider strategic investments for growth, such as R&D or market expansion, while maintaining current efficiency levels."
            elif health_score >= 60:
                return f"For {company_name}, focus on improving profitability and extending the cash runway for better resilience against market fluctuations. Operational efficiency is key."
            else:
                return f"It's crucial for {company_name} to prioritize cash flow management, implement aggressive cost optimization, and actively seek additional funding or strategic partnerships to stabilize the financial position."
        
        elif 'comparison' in query_lower or 'compare' in query_lower:
            return "I can help compare your company's metrics against industry benchmarks or historical performance if multiple reports are available. What specific metrics or periods would you like to compare?"
        
        elif 'trend' in query_lower or 'forecast' in query_lower or 'future' in query_lower:
            if predictions and predictions.get('revenue') and predictions['revenue'].get('trend'):
                revenue_trend = predictions['revenue']['trend']
                profit_trend = predictions['profit']['trend']
                cash_trend = predictions['cash']['trend']
                return f"The realistic forecast for {company_name} indicates an **{revenue_trend}** revenue trend, **{profit_trend}** profit trend, and **{cash_trend}** cash flow trend over the next 5 years. You can view detailed charts in the 'Predictions' tab."
            else:
                return "I need to run the full analysis to provide future predictions. Please ensure a report is processed."

        elif 'kpi' in query_lower or 'kpis' in query_lower:
            kpis = analysis.get('kpis', {})
            if kpis:
                top_kpis = [f"{k.replace('_', ' ').title()}: {v:.2f}" for k,v in list(kpis.items())[:5]]
                return f"Some key KPIs for {company_name} are: {', '.join(top_kpis)}. You can find a comprehensive list in the 'Dashboard' and 'Analytics' tabs."
            else:
                return "KPIs are not yet calculated. Please ensure a report has been processed."
        
        else:
            return f"I'm here to help {company_name} with financial analysis. Ask me about revenue, profitability, cash flow, risks, health score, or get recommendations for your business."

 
def generate_chart_image(fig, filename):
    """Saves a Plotly figure to an image file."""
    temp_folder = 'temp_charts'
    os.makedirs(temp_folder, exist_ok=True)
    path = os.path.join(temp_folder, filename)
    fig.write_image(path, scale=2)  
    return path

def create_advanced_pdf_report(report_data, logo_path=None):
        if not REPORTLAB_AVAILABLE:
            raise ImportError("ReportLab library not found, cannot generate PDF.")

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter,
                                rightMargin=inch/2, leftMargin=inch/2,
                                topMargin=inch/2, bottomMargin=inch/2)
        
        styles = getSampleStyleSheet()
        story = []

        # --- Custom Styles ---
        h1_style = ParagraphStyle(name='H1', parent=styles['h1'], alignment=TA_CENTER)
        h2_style = ParagraphStyle(name='H2', parent=styles['h2'], spaceBefore=20, spaceAfter=10)
        normal_style = styles['Normal']

        # --- Cover Page ---
        company_name = report_data['metrics'].get('company_name', 'N/A')
        if logo_path:
            img = Image(logo_path, width=1.5*inch, height=1.5*inch)
            img.hAlign = 'CENTER'
            story.append(img)
            story.append(Spacer(1, 0.3 * inch))

        story.append(Paragraph("Confidential Financial Analysis", h1_style))
        story.append(Spacer(1, 0.2 * inch))
        story.append(Paragraph(f"For: {company_name}", styles['h2']))
        story.append(Paragraph(f"Period: {report_data['metrics'].get('financial_year', 'N/A')}", styles['h3']))
        story.append(Paragraph(f"Report Generated: {datetime.now().strftime('%Y-%m-%d')}", styles['Normal']))
        story.append(PageBreak())

        # --- Executive Summary ---
        story.append(Paragraph("Executive Summary", h2_style))
        executive_summary_html = report_data.get('report', 'No summary available.').replace('**', '<b>').replace('\n', '<br/>')
        story.append(Paragraph(executive_summary_html, normal_style))
        story.append(Spacer(1, 0.3 * inch))

        # --- Key Metrics Table ---
        story.append(Paragraph("Key Financial Metrics", h2_style))
        metrics = report_data['metrics']
        analysis = report_data['analysis']
        metrics_data = [
            ['Metric', 'Value', 'Metric', 'Value'],
            ['Revenue (Cr)', f"₹{metrics.get('revenue', 0):.2f}", 'Health Score', f"{analysis.get('health_score', 0):.0f}/100"],
            ['Net Profit (Cr)', f"₹{metrics.get('profit', 0):.2f}", 'Risk Level', analysis.get('risk_level', 'N/A').upper()],
            ['Profitability', f"{metrics.get('profitability', 0):.1f}%", 'Cash Runway', f"{metrics.get('runway_months', 0):.1f} Months"],
        ]
        
        t = Table(metrics_data, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c3e50')), ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,0), 12), ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#f0f2f5')),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(t)
        story.append(Spacer(1, 0.3 * inch))

        # --- Generate and Add Charts ---
        story.append(Paragraph("Visual Analysis", h2_style))
        
        fig1 = px.bar(
            x=['Revenue', 'Profit', 'Assets', 'Liabilities', 'Cash'],
            y=[metrics.get('revenue',0), metrics.get('profit',0), metrics.get('assets',0), metrics.get('liabilities',0), metrics.get('cash',0)],
            title="Financial Overview (in Crores)"
        )
        chart1_path = generate_chart_image(fig1, 'overview.png')
        story.append(Image(chart1_path, width=7*inch, height=3.5*inch))
        story.append(PageBreak())

        fig2 = go.Figure(go.Indicator(
            mode = "gauge+number", value = analysis.get('health_score', 0),
            title = {'text': "Financial Health Score"},
            gauge = {'axis': {'range': [None, 100]}, 'bar': {'color': "#2c3e50"},
                    'steps' : [{'range': [0, 50], 'color': "#e74c3c"}, {'range': [50, 80], 'color': "#f39c12"}, {'range': [80, 100], 'color': "#27ae60"}]}))
        chart2_path = generate_chart_image(fig2, 'gauge.png')
        
        risks = analysis.get('risks', [])
        risk_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
        for risk in risks: risk_counts[risk.get('level', 'low').capitalize()] += 1
        fig3 = px.pie(values=list(risk_counts.values()), names=list(risk_counts.keys()), title="Risk Distribution", hole=.3,
                    color_discrete_map={'Critical':'#e74c3c', 'High':'#f39c12', 'Medium':'#3498db', 'Low':'#27ae60'})
        chart3_path = generate_chart_image(fig3, 'pie.png')

        chart_table = Table([[Image(chart2_path, width=3.5*inch, height=2.5*inch), Image(chart3_path, width=3.5*inch, height=2.5*inch)]])
        story.append(chart_table)

        # --- Build the PDF ---
        def add_page_number(canvas, doc):
            page_num = canvas.getPageNumber()
            text = f"Page {page_num} | {company_name} - Confidential Report"
            canvas.drawCentredString(letter[0]/2, 0.5 * inch, text)

        doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
        
        # --- Clean up temporary files ---
        for path in [chart1_path, chart2_path, chart3_path]: os.remove(path)
        if logo_path and 'temp_logo' in logo_path: os.remove(logo_path)

        output.seek(0)
        return output

def create_excel_report(report_data):
    """Create comprehensive Excel report with multiple sheets."""
    output = io.BytesIO()

    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            company_name = report_data['metrics'].get('company_name', 'Report').replace(' ', '_')
            financial_year = report_data['metrics'].get('financial_year', 'N/A')

            # --- 1. Summary Sheet ---
            summary_df_data = {
                'Metric': [
                    'Company Name', 'Financial Year', 'Health Score', 'Risk Level',
                    'Revenue (Cr)', 'Net Profit (Cr)', 'Profitability (%)',
                    'Cash (Cr)', 'Runway (Months)', 'Monthly Burn (Cr)',
                    'Current Ratio', 'Debt-to-Equity'
                ],
                'Value': [
                    report_data['metrics'].get('company_name', 'N/A'),
                    financial_year,
                    f"{report_data['analysis'].get('health_score', 0):.0f}/100",
                    report_data['analysis'].get('risk_level', 'unknown').upper(),
                    f"₹{report_data['metrics'].get('revenue', 0):.2f}",
                    f"₹{report_data['metrics'].get('profit', 0):.2f}",
                    f"{report_data['metrics'].get('profitability', 0):.2f}%",
                    f"₹{report_data['metrics'].get('cash', 0):.2f}",
                    f"{report_data['metrics'].get('runway_months', 0):.1f}",
                    f"₹{report_data['metrics'].get('burn_rate', 0):.2f}",
                    f"{report_data['metrics'].get('current_ratio', 0):.2f}",
                    f"{report_data['metrics'].get('debt_to_equity', 0):.2f}"
                ]
            }
            pd.DataFrame(summary_df_data).to_excel(writer, sheet_name='Summary', index=False)

            # --- 2. Financial Metrics Sheet ---
            metrics_df = pd.DataFrame([report_data['metrics']]).T
            metrics_df.columns = ['Value']
            metrics_df.index.name = 'Metric'
            metrics_df.to_excel(writer, sheet_name='Financial_Metrics')

            # --- 3. KPIs Sheet ---
            kpis_df = pd.DataFrame([report_data['analysis']['kpis']]).T
            kpis_df.columns = ['Value']
            kpis_df.index.name = 'KPI'
            kpis_df.to_excel(writer, sheet_name='KPIs')

            # --- 4. Risks Sheet ---
            if report_data['analysis']['risks']:
                risks_df = pd.DataFrame(report_data['analysis']['risks'])
                risks_df.to_excel(writer, sheet_name='Risk_Analysis', index=False)
            else:
                pd.DataFrame([{"Message": "No significant risks identified."}]).to_excel(writer, sheet_name='Risk_Analysis', index=False)
            
            # --- 5. Opportunities Sheet ---
            if report_data['analysis']['opportunities']:
                opportunities_df = pd.DataFrame(report_data['analysis']['opportunities'])
                opportunities_df.to_excel(writer, sheet_name='Opportunities', index=False)
            else:
                pd.DataFrame([{"Message": "No significant opportunities identified."}]).to_excel(writer, sheet_name='Opportunities', index=False)

            # --- 6. Predictions Sheet (Realistic, Conservative, Optimistic) ---
            predictions_data = []
            future_months_count = 12 * 5
            
            for scenario in ['conservative', 'realistic', 'optimistic']:
                for metric_key_prefix in ['revenue', 'profit', 'cash']:
                    pred_key = f'{metric_key_prefix}_{scenario}'
                    if pred_key in report_data['predictions']:
                        values = report_data['predictions'][pred_key]['values'][:future_months_count]
                        trend = report_data['predictions'][pred_key]['trend']

                        for month_idx in range(len(values)):
                            predictions_data.append({
                                'Month': month_idx + 1,
                                'Scenario': scenario.title(),
                                'Metric': metric_key_prefix.title(),
                                'Value (Cr)': values[month_idx],
                                'Trend': trend
                            })

            if predictions_data:
                pred_df = pd.DataFrame(predictions_data)
                pred_df.to_excel(writer, sheet_name='Predictions', index=False)
            else:
                pd.DataFrame([{"Message": "No predictions generated."}]).to_excel(writer, sheet_name='Predictions', index=False)

            # --- 7. Executive Summary Text Sheet ---
            executive_summary_text = report_data.get('report', 'No executive summary available.')
            summary_text_df = pd.DataFrame({'Report Content': [executive_summary_text]})
            summary_text_df.to_excel(writer, sheet_name='Executive_Summary_Text', index=False)

        output.seek(0)
        return output
    except Exception as e:
        logging.error(f"Error creating Excel report: {e}", exc_info=True)
        raise

def create_pdf_report(report_data):
    """Create comprehensive PDF report using ReportLab."""
    if not Paragraph:
        raise ImportError("ReportLab library not available for PDF generation.")

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=letter)
    styles = getSampleStyleSheet()
    
    h1_style = styles['h1']
    h1_style.spaceAfter = 14
    h2_style = styles['h2']
    h2_style.spaceAfter = 12
    h3_style = styles['h3']
    h3_style.spaceAfter = 10
    normal_style = styles['Normal']
    normal_style.fontSize = 10
    normal_style.leading = 14
    bold_style = ParagraphStyle(name='BoldNormal', parent=normal_style, fontName='Helvetica-Bold')
    bullet_style = ParagraphStyle(name='Bullet', parent=normal_style, leftIndent=36, bulletIndent=18, bulletFontSize=10)
    
    story = []
#  --- Header (No changes needed) ---
    story.append(Paragraph("AI CFO Assistant: Financial Analysis Report", h1_style))
    story.append(Spacer(1, 0.2 * inch))

    company_name = report_data['metrics'].get('company_name', 'Unknown Company')
    financial_year = report_data['metrics'].get('financial_year', 'Unknown Year')
    story.append(Paragraph(f"Company: <font color='blue'><b>{company_name}</b></font>", normal_style))
    story.append(Paragraph(f"Financial Year: <b>{financial_year}</b>", normal_style))
    story.append(Spacer(1, 0.2 * inch))

    story.append(Paragraph("Executive Summary", h2_style))
    executive_summary_html = report_data.get('report', 'No executive summary available.').replace('**', '<b>').replace('\n', '<br/>')
    story.append(Paragraph(executive_summary_html, normal_style))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Key Financial Metrics", h2_style))
    metrics_data = [
        ['Metric', 'Value'],
        ['Revenue (Crores)', f"₹{report_data['metrics'].get('revenue', 0):.2f}"],
        ['Net Profit (Crores)', f"₹{report_data['metrics'].get('profit', 0):.2f}"],
        ['Profitability (%)', f"{report_data['metrics'].get('profitability', 0):.2f}%"],
        ['Cash Reserves (Crores)', f"₹{report_data['metrics'].get('cash', 0):.2f}"],
        ['Runway (Months)', f"{report_data['metrics'].get('runway_months', 0):.1f}"],
        ['Health Score', f"{report_data['analysis'].get('health_score', 0):.0f}/100"],
        ['Risk Level', report_data['analysis'].get('risk_level', 'unknown').upper()],
        ['Current Ratio', f"{report_data['metrics'].get('current_ratio', 0):.2f}"],
        ['Debt-to-Equity', f"{report_data['metrics'].get('debt_to_equity', 0):.2f}"]
    ]
    metrics_table = Table(metrics_data, colWidths=[2.5*inch, 3*inch])
    metrics_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 12),
        ('BACKGROUND', (0,1), (-1,-1), colors.white),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#dee2e6')),
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))
    story.append(metrics_table)
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Risk Assessment", h2_style))
    risks = report_data.get('analysis', {}).get('risks', [])
    if risks:
        for risk in risks:
            story.append(Paragraph(f"<b>• {risk.get('level').upper()} Risk ({risk['type'].title()}):</b> {risk.get('message')}", bullet_style))
            story.append(Paragraph(f"<b>Impact:</b> {risk.get('impact')}", normal_style, leftIndent=54))
            story.append(Paragraph(f"<b>Recommendation:</b> {risk.get('recommendation')}", normal_style, leftIndent=54))
            story.append(Paragraph(f"<b>Timeline:</b> {risk.get('timeline')}", normal_style, leftIndent=54))
            story.append(Paragraph(f"<b>KPIs Affected:</b> {', '.join(risk.get('kpi_affected', ['N/A']))}<br/>", normal_style, leftIndent=54))
            story.append(Spacer(1, 0.1 * inch))
    else:
        story.append(Paragraph("No significant risks identified.", normal_style))
    story.append(Spacer(1, 0.3 * inch))

    story.append(Paragraph("Strategic Opportunities", h2_style))
    opportunities = report_data.get('analysis', {}).get('opportunities', [])
    if opportunities:
        for opp in opportunities:
            story.append(Paragraph(f"<b>• Opportunity ({opp.get('type').title()}):</b> {opp.get('message')}", bullet_style))
            story.append(Paragraph(f"<b>Impact:</b> {opp['impact']}", normal_style, leftIndent=54))
            story.append(Paragraph(f"<b>Recommendation:</b> {opp.get('recommendation')}", normal_style, leftIndent=54))
            story.append(Paragraph(f"<b>Timeline:</b> {opp.get('timeline')}<br/>", normal_style, leftIndent=54))
            story.append(Spacer(1, 0.1 * inch))
    else:
        story.append(Paragraph("No explicit opportunities identified.", normal_style))
    story.append(Spacer(1, 0.3 * inch))

    doc.build(story)
    output.seek(0)
    return output

# --- Agent Initialization ---
ingestion_agent = AdvancedIngestionAgent()
analysis_agent = AdvancedAnalysisAgent()
advisor_agent = AdvancedAdvisorAgent()

# --- Flask Routes ---

 

@app.route('/')
def index():
    user = None
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
    return render_template('index.html', user=user)

@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'POST':
        fullname = request.form['fullname']
        email = request.form['email']
        password = request.form['password']
 
        if email in users:
            flash("user already exist",'danger')
            return redirect(url_for('signup'))
        
        users[email]={'fullname':fullname, 'password':password}

        flash('Sign up successful','success')
        return redirect(url_for('login'))

    return render_template('signup.html')

@app.route('/signin', methods=['GET', 'POST'])

def login():
    """Handles user sign-in using the in-memory dictionary."""
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')
        
        user = users.get(email)

        # Check if user exists and if the hashed password matches
        if user and user['password']==password:
            # Store the user's info in the session
            session['user'] = {'email': email, 'fullname': user['fullname']}
            flash(f"Welcome back, {user['fullname']}!", "success")
            return redirect(url_for('dashboard'))
        else:
            flash("Invalid email or password. Please try again.", "danger")
            # Redirect back to the signin page on failure
            return redirect(url_for('login'))

    return render_template('login.html')

@app.route('/logout')
def logout():
    session.pop('user_id', None)
    session.pop('user', None)
    session.pop('reports', None)
    session.pop('chat_history', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    
    """Displays the main application dashboard for logged-in users."""
    return render_template('dashboard.html')

@app.route('/user_info', methods=['GET'])
def get_user_info():
    if 'user_id' in session:
        user = User.query.get(session['user_id'])
        if user:
            return jsonify({
                'logged_in': True,
                'username': user.username,
                'company_name': user.company_name
            })
    return jsonify({'logged_in': False})

@app.route('/upload', methods=['POST'])
# @login_required
def upload_files():
    try:
        files = request.files.getlist('files')
        if not files:
            return jsonify({'success': False, 'error': 'No files provided'}), 400
        
        if 'reports' not in session:
            session['reports'] = []

        results = []
        for file in files:
            if file and allowed_file(file.filename):
                original_name = secure_filename(file.filename)
                unique_filename = f"{datetime.now().strftime('%Y%m%d%H%M%S%f')}_{original_name}"
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_filename)
                file.save(file_path)
                logging.info(f"File saved temporarily: {file_path}")

                try:
                    logging.info(f"Starting processing of {original_name}")
                    metrics, tables, extracted_raw_data = ingestion_agent.process_file(file_path)
                    
                    analysis = analysis_agent.analyze_comprehensive_health(metrics)
                    predictions = analysis_agent.predict_advanced_trends(metrics)

                    result = sanitize_for_json({
                        'id': str(len(session['reports']) + len(results)),
                        'file_name': original_name,
                        'original_file_path': file_path,
                        'metrics': metrics,
                        'analysis': analysis,
                        'predictions': predictions,
                        'processing_time': metrics.get('processing_time', 0),
                        'confidence_score': metrics.get('confidence_score', 0),
                        'upload_time': datetime.now().isoformat(),
                        'status': 'processed'
                    })
                    results.append(result)
                    logging.info(f"Successfully processed {original_name} in {metrics.get('processing_time', 0):.0f}ms")

                except Exception as processing_error:
                    logging.error(f"Processing failed for {original_name}: {processing_error}", exc_info=True)
                    results.append({
                        'id': str(len(session['reports']) + len(results)),
                        'file_name': original_name,
                        'error': str(processing_error),
                        'status': 'failed',
                        'upload_time': datetime.now().isoformat()
                    })
                finally:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                        logging.info(f"Removed temporary file: {file_path}")
            else:
                results.append({
                    'id': str(len(session['reports']) + len(results)),
                    'file_name': file.filename,
                    'error': f'File {file.filename} is not allowed or invalid.',
                    'status': 'failed',
                    'upload_time': datetime.now().isoformat()
                })

        if not results:
            return jsonify({'success': False, 'error': 'No files were successfully processed'}), 500
        
        session['reports'].extend(results)
        
        company_relationships = {}
        if len(session['reports']) > 0:
            successful_reports_in_session = [r for r in session['reports'] if r.get('status') == 'processed']
            
            if len(successful_reports_in_session) > 0: 
                company_relationships = analysis_agent.detect_company_relationships(successful_reports_in_session)
                if company_relationships:
                    for i, report in enumerate(session['reports']):
                        if report.get('status') == 'processed':
                            for company_group_name, related_original_indices in company_relationships.items():
                                if i in related_original_indices:
                                    related_successful_reports = [session['reports'][idx] for idx in related_original_indices if session['reports'][idx].get('status') == 'processed']
                                    
                                    if len(related_successful_reports) >= 2:
                                        comparison_data = analysis_agent.compare_reports(related_successful_reports)
                                        session['reports'][i]['comparison_data'] = comparison_data
                                        session['reports'][i]['related_reports_count'] = len(related_successful_reports)
                                        session['reports'][i]['company_group'] = company_group_name
                                    else:
                                        session['reports'][i]['related_reports_count'] = 1
                                        session['reports'][i]['company_group'] = report['metrics'].get('company_name', 'N/A')
                                    break
                            if 'company_group' not in session['reports'][i]:
                                session['reports'][i]['related_reports_count'] = 1
                                session['reports'][i]['company_group'] = report['metrics'].get('company_name', 'N/A')
                        else:
                            session['reports'][i]['related_reports_count'] = 1
                            session['reports'][i]['company_group'] = 'N/A (Failed)'

        session.modified = True
        logging.info(f"Total reports in session: {len(session['reports'])}")

        for res in results:
            if res.get('status') != 'failed':
                comp_for_report = None
                for session_report in session['reports']:
                    if session_report.get('id') == res.get('id'):
                        comp_for_report = session_report.get('comparison_data')
                        break
                res['report'] = advisor_agent.generate_executive_report(res['metrics'], res['analysis'], res['predictions'], comp_for_report)
        
        return jsonify({
            'success': True,
            'results': results,
            'message': f'Successfully processed {len(results)} file(s)',
            'total_reports_in_session': len(session['reports'])
        })

    except Exception as e:
        logging.error(f"Upload processing failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'Server error during upload: {str(e)}'}), 500

@app.route('/generate_charts', methods=['POST'])
# @login_required
def generate_charts():
    try:
        data = request.get_json()
        report_id = data.get('report_id')
        
        reports = session.get('reports', [])
        target_report = None
        for r in reports:
            if r.get('id') == report_id:
                target_report = r
                break
        
        if not target_report:
            if reports:
                target_report = reports[-1]
            else:
                return jsonify({'success': False, 'error': 'No reports available or invalid report ID.'}), 400

        if target_report.get('status') == 'failed' or 'metrics' not in target_report or 'analysis' not in target_report:
            return jsonify({'success': False, 'error': 'Selected report is incomplete or failed to process.'}), 400

        metrics = target_report['metrics']
        analysis = target_report['analysis']
        predictions = target_report['predictions']
        company_comparison = target_report.get('comparison_data', {})

        charts_json = {} # Will store Plotly JSON strings

        # --- 1. Financial Overview Bar Chart ---
        try:
            fig1 = px.bar(
                x=['Revenue', 'Expenses', 'Profit', 'Assets', 'Liabilities', 'Cash'],
                y=[(metrics.get('revenue',0) or 0), (metrics.get('expenses',0) or 0), (metrics.get('profit',0) or 0),
                   (metrics.get('assets',0) or 0), (metrics.get('liabilities',0) or 0), (metrics.get('cash',0) or 0)],
                title=f"{metrics.get('company_name', 'Company')} - Financial Overview (in Crores)",
                labels={'x': 'Category', 'y': 'Amount (Crores)'},
                color_discrete_sequence=['#2E86AB', '#A23B72', '#F18F01', '#5B84B1', '#FC766A', '#3BB27B']
            )
            charts_json['financial_overview'] = fig1.to_json()
        except Exception as e:
            logging.error(f"Error generating financial_overview chart: {e}")
            charts_json['financial_overview'] = json.dumps({'error': f'Chart generation failed: {str(e)}'})

        # --- 2. KPI Comparison Chart ---
        try:
            kpi_data = {
                'metrics': [],
                'actual_values': [],
                'benchmark_good': [],
                'benchmark_average': []
            }
            
            kpi_labels_and_keys = [
                ('Net Profit Margin (%)', 'net_profit_margin', 'profitability'),
                ('Current Ratio', 'current_ratio', 'current_ratio'),
                ('ROA (%)', 'roa', 'roa'),
                ('ROE (%)', 'roe', 'roe'),
                ('Debt/Equity', 'debt_to_equity', 'debt_to_equity'),
                ('Cash Ratio', 'cash_ratio', 'cash_ratio')
            ]
            
            for label, kpi_key, benchmark_key in kpi_labels_and_keys:
                actual_value = (analysis['kpis'].get(kpi_key, 0.0) or 0)
                
                kpi_data['metrics'].append(label)
                kpi_data['actual_values'].append(actual_value)
                
                if benchmark_key in analysis['benchmarks']:
                    kpi_data['benchmark_good'].append(analysis['benchmarks'][benchmark_key]['good'] or 0)
                    kpi_data['benchmark_average'].append(analysis['benchmarks'][benchmark_key]['average'] or 0)
                else:
                    kpi_data['benchmark_good'].append(0)
                    kpi_data['benchmark_average'].append(0)

            fig2 = go.Figure(data=[
                go.Bar(name='Actual', x=kpi_data['metrics'], y=kpi_data['actual_values'], marker_color='#3498db'),
                go.Bar(name='Good Benchmark', x=kpi_data['metrics'], y=kpi_data['benchmark_good'], marker_color='#2ecc71', opacity=0.7),
                go.Bar(name='Average Benchmark', x=kpi_data['metrics'], y=kpi_data['benchmark_average'], marker_color='#f39c12', opacity=0.7)
            ])
            fig2.update_layout(barmode='group', title=f"{metrics.get('company_name', 'Company')} - KPI Performance vs Benchmarks")
            charts_json['kpi_comparison'] = fig2.to_json()
        except Exception as e:
            logging.error(f"Error generating kpi_comparison chart: {e}")
            charts_json['kpi_comparison'] = json.dumps({'error': f'Chart generation failed: {str(e)}'})

        # --- 3. Health Score Gauge ---
        try:
            fig3 = go.Figure(go.Indicator(
                mode = "gauge+number+delta",
                value = (analysis.get('health_score', 0) or 0),
                domain = {'x': [0, 1], 'y': [0, 1]},
                title = {'text': f"{metrics.get('company_name', 'Company')} - Financial Health Score"},
                delta = {'reference': 80, 'increasing': {'color': "#2ecc71"}, 'decreasing': {'color': "#e74c3c"}},
                gauge = {'axis': {'range': [None, 100], 'tickwidth': 1, 'tickcolor': "darkblue"},
                         'bar': {'color': "darkblue"},
                         'steps': [
                             {'range': [0, 50], 'color': "#e74c3c"},
                             {'range': [50, 70], 'color': "#f39c12"},
                             {'range': [70, 85], 'color': "#3498db"},
                             {'range': [85, 100], 'color': "#2ecc71"}],
                         'threshold': {'line': {'color': "red", 'width': 4},
                                       'thickness': 0.75, 'value': 90}}))
            charts_json['health_gauge'] = fig3.to_json()
        except Exception as e:
            logging.error(f"Error generating health_gauge chart: {e}")
            charts_json['health_gauge'] = json.dumps({'error': f'Chart generation failed: {str(e)}'})

        # --- 4. Multi-scenario Revenue Prediction ---
        try:
            forecast_months_count = len(predictions['revenue_realistic']['values'])
            months = [f'Month {i}' for i in range(1, forecast_months_count + 1)]

            fig4 = go.Figure()
            fig4.add_trace(go.Scatter(x=months, y=predictions['revenue_conservative']['values'],
                                      mode='lines', name='Conservative', line=dict(color='#95a5a6', dash='dot')))
            fig4.add_trace(go.Scatter(x=months, y=predictions['revenue_realistic']['values'],
                                      mode='lines+markers', name='Realistic', line=dict(color='#3498db')))
            fig4.add_trace(go.Scatter(x=months, y=predictions['revenue_optimistic']['values'],
                                      mode='lines', name='Optimistic', line=dict(color='#2ecc71', dash='dash')))
            fig4.update_layout(title=f"{metrics.get('company_name', 'Company')} - Revenue Forecast (Multi-Scenario, Next 5 Years)",
                               xaxis_title='Time Period', yaxis_title='Revenue (Crores)')
            charts_json['revenue_scenarios'] = fig4.to_json()
        except Exception as e:
            logging.error(f"Error generating revenue_scenarios chart: {e}")
            charts_json['revenue_scenarios'] = json.dumps({'error': f'Chart generation failed: {str(e)}'})

        # --- 5. Cash Flow Projection ---
        try:
            fig5 = go.Figure()
            fig5.add_trace(go.Scatter(x=months, y=predictions['cash_realistic']['values'],
                                      mode='lines+markers', name='Cash Balance',
                                      line=dict(color='#1abc9c'), fill='tozeroy', fillcolor='rgba(26, 188, 156, 0.2)'))
            fig5.update_layout(title=f"{metrics.get('company_name', 'Company')} - Cash Flow Projection (Next 5 Years)",
                               xaxis_title='Time Period', yaxis_title='Cash (Crores)')
            charts_json['cash_flow_projection'] = fig5.to_json()
        except Exception as e:
            logging.error(f"Error generating cash_flow_projection chart: {e}")
            charts_json['cash_flow_projection'] = json.dumps({'error': f'Chart generation failed: {str(e)}'})
        
        # --- 6. Profitability Trend ---
        try:
            predicted_profitability = []
            for i in range(len(predictions['revenue_realistic']['values'])):
                rev = (predictions['revenue_realistic']['values'][i] or 0)
                prof = (predictions['profit_realistic']['values'][i] or 0)
                if rev > 0:
                    predicted_profitability.append((prof / rev) * 100)
                else:
                    predicted_profitability.append(0)

            fig6 = go.Figure()
            fig6.add_trace(go.Scatter(x=months, y=predicted_profitability,
                                      mode='lines+markers', name='Profitability Margin',
                                      line=dict(color='#e67e22')))
            fig6.update_layout(title=f"{metrics.get('company_name', 'Company')} - Profitability Trend Forecast (Next 5 Years)",
                               xaxis_title='Time Period', yaxis_title='Profitability (%)')
            charts_json['profitability_trend'] = fig6.to_json()
        except Exception as e:
            logging.error(f"Error generating profitability_trend chart: {e}")
            charts_json['profitability_trend'] = json.dumps({'error': f'Chart generation failed: {str(e)}'})

        # --- 7. Risk Distribution Pie Chart ---
        try:
            risks = analysis.get('risks', [])
            risk_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
            for risk in risks:
                level = risk.get('level', 'Low').capitalize()
                if level in risk_counts:
                    risk_counts[level] += 1
            
            fig7 = px.pie(values=list(risk_counts.values()), names=list(risk_counts.keys()),
                          title=f"{metrics.get('company_name', 'Company')} - Risk Distribution by Severity",
                          color_discrete_map={'Critical':'#e74c3c', 'High':'#f39c12', 'Medium':'#3498db', 'Low':'#2ecc71'})
            charts_json['risk_distribution'] = fig7.to_json()
        except Exception as e:
            logging.error(f"Error generating risk_distribution chart: {e}")
            charts_json['risk_distribution'] = json.dumps({'error': f'Chart generation failed: {str(e)}'})

        # --- 8. Company Comparison Chart (if multiple reports for same company are present) ---
        try:
            if company_comparison and company_comparison.get('metrics_over_time'):
                comp_years = sorted([int(y) for y in company_comparison['metrics_over_time'].keys()])
                if len(comp_years) >= 2:
                    comp_revenue = [(company_comparison['metrics_over_time'][str(y)]['revenue'] or 0) for y in comp_years]
                    comp_profit = [(company_comparison['metrics_over_time'][str(y)]['profit'] or 0) for y in comp_years]
                    comp_health_score = [(company_comparison['metrics_over_time'][str(y)]['health_score'] or 0) for y in comp_years]
                    
                    fig8 = go.Figure()
                    fig8.add_trace(go.Bar(x=comp_years, y=comp_revenue, name='Revenue', marker_color='#3498db'))
                    fig8.add_trace(go.Bar(x=comp_years, y=comp_profit, name='Profit', marker_color='#2ecc71'))
                    fig8.add_trace(go.Scatter(x=comp_years, y=comp_health_score, mode='lines+markers', name='Health Score', yaxis='y2', line=dict(color='#e74c3c', width=3, dash='dot')))
                    
                    fig8.update_layout(
                        title=f"{metrics.get('company_name', 'Company')} - Year-over-Year Comparison",
                        xaxis_title='Financial Year',
                        yaxis=dict(title='Amount (Crores)', side='left'),
                        yaxis2=dict(title='Health Score', overlaying='y', side='right', range=[0,100], showgrid=False),
                        barmode='group'
                    )
                    charts_json['company_comparison'] = fig8.to_json()
                else:
                    charts_json['company_comparison'] = ""
            else:
                charts_json['company_comparison'] = ""
        except Exception as e:
            logging.error(f"Error generating company_comparison chart: {e}")
            charts_json['company_comparison'] = json.dumps({'error': f'Chart generation failed: {str(e)}'})


        return jsonify({
            'success': True,
            'charts': charts_json,
            'summary': {
                'total_charts': len(charts_json),
                'health_score': analysis.get('health_score', 0),
                'risk_level': analysis.get('risk_level', 'unknown'),
                'confidence_score': metrics.get('confidence_score', 0)
            }
        })
    except Exception as e:
        logging.error(f"Overall chart generation route failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'Chart generation failed: {str(e)}'}), 500

# --- REPLACE your old /export_report route with this correct one ---
@app.route('/export_report', methods=['POST'])
# @login_required
def export_report():
    try:
        # This logic correctly handles both JSON data and FormData for files
        if 'report_id' in request.form:
            data = request.form
        else:
            data = request.get_json()

        report_id = data.get('report_id')
        format_type = data.get('format')
        
        target_report = next((r for r in session.get('reports', []) if r.get('id') == report_id), None)

        if not target_report:
            return jsonify({'success': False, 'error': 'Report not found.'}), 404

        company_name_safe = secure_filename(target_report['metrics'].get('company_name', 'report'))
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        if format_type == 'pdf':
            logo_path = None
            if 'logo' in request.files:
                logo_file = request.files['logo']
                if logo_file and allowed_file(logo_file.filename):
                    logo_filename = f"temp_logo_{secure_filename(logo_file.filename)}"
                    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
                    logo_file.save(logo_path)
            
            # This now calls your new ADVANCED function
            pdf_output = create_advanced_pdf_report(target_report, logo_path)
            filename = f"financial_report_{company_name_safe}_{timestamp}.pdf"
            
            reports_path = os.path.join(app.config['REPORTS_FOLDER'], filename)
            with open(reports_path, 'wb') as f:
                f.write(pdf_output.getvalue())

            return jsonify({'success': True, 'filename': filename})
        
        elif format_type == 'json':
            # (Handling for other formats like JSON)
            filename = f"financial_report_{company_name_safe}_{timestamp}.json"
            return jsonify({'success': True, 'data': sanitize_for_json(target_report), 'filename': filename})
        
        elif format_type == 'excel':
             excel_output = create_excel_report(target_report)
             filename = f"financial_analysis_{company_name_safe}_{timestamp}.xlsx"
             reports_path = os.path.join(app.config['REPORTS_FOLDER'], filename)
             with open(reports_path, 'wb') as f:
                f.write(excel_output.getvalue())
             return jsonify({'success': True, 'filename': filename})

    except Exception as e:
        logging.error(f"Export failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'Export failed: {str(e)}'}), 500

@app.route('/download_report/<filename>')
# @login_required
def download_report(filename):
    """Download generated report file."""
    try:
        base = os.path.basename(filename)
        return send_file(
            os.path.join(app.config['REPORTS_FOLDER'], base),
            as_attachment=True,
            download_name=base
        )
    except FileNotFoundError:
        return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        logging.error(f"Download failed: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500

@app.route('/get_reports', methods=['GET'])
# @login_required
def get_reports():
    """Get list of stored (processed) reports from session."""
    # This version correctly returns the FULL report object from the session.
    all_reports = session.get('reports', [])
    all_reports.sort(key=lambda x: x.get('upload_time', ''), reverse=True)
    
    return jsonify({
        'success': True,
        'reports': all_reports,
        'total_files': len(all_reports)
    })

@app.route('/chat', methods=['POST'])
# @login_required
def chat():
    """
    Handles chat requests with a streaming response to support the live "typing"
    effect on the frontend.
    """
    try:
        data = request.get_json()
        query = data.get('query', '').strip()
        report_id = data.get('report_id')

        if not query:
            # For non-streaming errors, jsonify is fine as the JS expects it on failure.
            return jsonify({'success': False, 'error': 'Query cannot be empty'}), 400

        reports = session.get('reports', [])
        # A more concise way to find the target report
        target_report = next((r for r in reports if r.get('id') == report_id), None)
        
        # If no specific report is found or targeted, default to the latest one.
        if not target_report:
            target_report = reports[0] if reports else None
            if not target_report:
                # Return a plain text response for errors the stream handler will catch.
                return Response("Please upload a financial report first to start the analysis.", status=400)

        if target_report.get('status') == 'failed' or 'metrics' not in target_report:
            return Response("Cannot chat about a failed or incomplete report. Please select a valid one.", status=400)

        # Correct, concise way to initialize chat history.
        chat_history = session.get('chat_history', [])
        
        # This inner function is a generator. It will yield response chunks one by one.
        def generate():
            # This call returns a generator, not a complete string.
            response_generator = advisor_agent.chat_response(
                query,
                target_report['metrics'],
                target_report['analysis'],
                chat_history
            )
            
            full_response_text = ""
            for chunk in response_generator:
                full_response_text += chunk
                yield chunk  # Send each piece to the frontend as it arrives

            # --- This code runs ONLY AFTER the entire stream is complete ---
            # Now, update the session history with the complete query and response.
            chat_history.append({'role': 'user', 'content': query})
            chat_history.append({'role': 'assistant', 'content': full_response_text})
            
            # Keep the session history from growing too large.
            session['chat_history'] = chat_history[-20:]
            session.modified = True
        
        # Return the generator wrapped in a streaming Response object.
        # This is what enables the real-time effect.
        return Response(stream_with_context(generate()), mimetype='text/plain')

    except Exception as e:
        # Log the detailed error for debugging.
        logging.error(f"Chat streaming error: {e}", exc_info=True)
        # Return a simple, non-streaming error message for the frontend to display.
        return Response(f"An unexpected error occurred in the chat: {str(e)}", status=500)
    

# @app.route('/chat', methods=['POST'])
# # @login_required
# def chat():
#     try:
#         data = request.get_json()
#         query = data.get('query', '').strip()
#         report_id = data.get('report_id')

#         if not query:
#             return jsonify({'success': False, 'error': 'Query cannot be empty'}), 400

#         reports = session.get('reports', [])
#         target_report = None
#         for r in reports:
#             if r.get('id') == report_id:
#                 target_report = r
#                 break
        
#         # if not target_report:
#         #     if reports:
#         #         target_report = reports[-1]
#         #     else:
#         #         return jsonify({'success': False, 'response': 'Please upload a financial report first to start the analysis.'}), 400
#         if not target_report:
#             target_report = reports[-1] if reports else None
#             if not target_report:
#                 return Response("Please upload a financial report first to start the analysis.", status=400)
#         if target_report.get('status') == 'failed' or 'metrics' not in target_report:
#             return Response("Cannot chat about a failed or incomplete report.", status=400)

#         chat_history = session.get('chat_history', [])
        
#         response_text = advisor_agent.chat_response(
#             query,
#             target_report['metrics'],
#             target_report['analysis'],
#             session['chat_history']
#         )
        
#         session['chat_history'].append({'role': 'user', 'content': query})
#         session['chat_history'].append({'role': 'assistant', 'content': response_text})
        
#         if len(session['chat_history']) > 20:
#             session['chat_history'] = session['chat_history'][-20:]
        
#         session.modified = True
        
#         return jsonify({
#             'success': True,
#             'response': response_text,
#             'chat_history': session['chat_history']
#         })
#     except Exception as e:
#         logging.error(f"Chat error: {e}", exc_info=True)
#         return jsonify({'success': False, 'error': f'Chat processing failed: {str(e)}'}), 500

@app.route('/clear_session', methods=['POST'])
# @login_required
def clear_session():
    """Clear all session data for reports and chat history."""
    try:
        session.clear()
        logging.info("Session data cleared successfully.")
        return jsonify({'success': True, 'message': 'All data cleared successfully'})
    except Exception as e:
        logging.error(f"Error clearing session: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'Failed to clear session: {str(e)}'}), 500


@app.route('/demo', methods=['GET'])
# @login_required
def demo():
    """Generate demo data for testing."""
    try:
        demo_metrics = {
            'company_name': 'AI CFO Demo Corp',
            'financial_year': '2023-24',
            'revenue': 15000.0,
            'expenses': 9000.0,
            'profit': 6000.0,
            'assets': 25000.0,
            'liabilities': 5000.0,
            'cash': 8000.0,
            'equity': 20000.0,
            'processing_time': 1500.0,
            'file_name': 'Demo_Financial_Report.pdf'
        }
        
        demo_metrics['profitability'] = (demo_metrics['profit'] / demo_metrics['revenue'] * 100) if demo_metrics['revenue'] > 0 else 0
        demo_metrics['debt_to_equity'] = (demo_metrics['liabilities'] / demo_metrics['equity']) if demo_metrics['equity'] > 0 else 0
        demo_metrics['current_ratio'] = (demo_metrics['assets'] / demo_metrics['liabilities']) if demo_metrics['liabilities'] > 0 else 999.99
        monthly_expenses = demo_metrics['expenses'] / 12 if demo_metrics['expenses'] > 0 else 0
        demo_metrics['burn_rate'] = monthly_expenses
        demo_metrics['runway_months'] = (demo_metrics['cash'] / monthly_expenses) if monthly_expenses > 0 else 999.99
        demo_metrics['roa'] = (demo_metrics['profit'] / demo_metrics['assets'] * 100) if demo_metrics['assets'] > 0 else 0.0
        demo_metrics['roe'] = (demo_metrics['profit'] / demo_metrics['equity'] * 100) if demo_metrics['equity'] > 0 else 0.0
        demo_metrics['confidence_score'] = 95

        analysis = analysis_agent.analyze_comprehensive_health(demo_metrics)
        
        predictions = analysis_agent.predict_advanced_trends(demo_metrics)
        
        demo_result = sanitize_for_json({
            'id': 'demo_report_1',
            'file_name': 'Demo_Financial_Report_2023-24.pdf',
            'metrics': demo_metrics,
            'analysis': analysis,
            'predictions': predictions,
            'tables_extracted': [],
            'processing_time': demo_metrics['processing_time'],
            'confidence_score': demo_metrics['confidence_score'],
            'upload_time': datetime.now().isoformat(),
            'status': 'processed'
        })
        
        demo_result['report'] = advisor_agent.generate_executive_report(demo_metrics, analysis, predictions)


        if 'reports' not in session:
            session['reports'] = []
        
        demo_report_exists = False
        for i, r in enumerate(session['reports']):
            if r.get('id') == demo_result['id']:
                session['reports'][i] = demo_result
                demo_report_exists = True
                break
        
        if not demo_report_exists:
            session['reports'].append(demo_result)
        
        session.modified = True
        logging.info("Demo report generated successfully.")

        return jsonify({
            'success': True,
            'result': demo_result,
            'message': 'Demo report generated successfully',
            'total_reports_in_session': len(session['reports'])
        })
    except Exception as e:
        logging.error(f"Demo generation error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'Demo generation failed: {str(e)}'}), 500

@app.errorhandler(413)
def too_large(e):
    logging.error(f"Request entity too large: {e}")
    return jsonify({'success': False, 'error': 'File too large. Maximum size is 500MB'}), 413

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True, host='0.0.0.0', port=5000)

















 

# @app.route('/dashboard')
# @login_required  # This protects the dashboard
# def dashboard():
#     """Displays the main application dashboard for logged-in users."""
#     return render_template('dashboard.html')

# @app.route('/logout')
# def logout():
#     """Logs the user out."""
#     session.pop('user', None)
#     flash("You have been successfully logged out.", "success")
#     return redirect(url_for('home'))



# --- ADVANCED ROUTES FROM YOUR ORIGINAL APP.PY ---
# You can uncomment these and integrate them later.
# They are protected by @login_required.

# @app.route('/upload', methods=['POST'])
# @login_required
# def upload_files():
#     # Your full file upload logic here
#     return jsonify({'success': True, 'message': 'File upload is protected.'})

# @app.route('/generate_charts', methods=['POST'])
# @login_required
# def generate_charts():
#     # Your full chart generation logic here
#     return jsonify({'success': True, 'message': 'Chart generation is protected.'})

# @app.route('/chat', methods=['POST'])
# @login_required
# def chat():
#     # Your full chat logic here
#     return jsonify({'success': True, 'response': 'Chat is protected.'})

# --- Run the App ---
# if __name__ == '__main__':
#     app.run(debug=True)
